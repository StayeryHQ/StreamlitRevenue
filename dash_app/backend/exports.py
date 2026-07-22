# dash_app/backend/exports.py
# Shared Excel helpers for the per-page downloads. Pages call these from a
# dcc.Download callback via dcc.send_bytes(lambda buf: write_workbook(buf, ...)).

from __future__ import annotations

import pandas as pd
from openpyxl.utils import get_column_letter


def autosize(ws, max_width: int = 48) -> None:
    """Roughly fit column widths to content (openpyxl worksheet)."""
    for idx, col in enumerate(ws.iter_cols(values_only=True), start=1):
        width = max((len(str(v)) for v in col if v is not None), default=8)
        ws.column_dimensions[get_column_letter(idx)].width = min(width + 2, max_width)


def write_workbook(buf, sheets: dict[str, pd.DataFrame], *, index: bool = False) -> None:
    """Write {sheet_name: df} into an xlsx buffer; empty dict yields a note sheet."""
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        wrote = False
        for name, df in sheets.items():
            if df is None or df.empty:
                continue
            df.to_excel(writer, sheet_name=name[:31], index=index)
            autosize(writer.sheets[name[:31]])
            wrote = True
        if not wrote:
            pd.DataFrame({"Hinweis": ["Keine Daten im gewählten Zeitraum."]}).to_excel(
                writer, sheet_name="Hinweis", index=False)
