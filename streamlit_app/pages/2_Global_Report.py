"""Global Report"""

from __future__ import annotations

import sys
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(_REPO_ROOT / "src"))
sys.path.insert(0, str(_REPO_ROOT / "streamlit_app"))

import pandas as pd
import streamlit as st

from components import cached_data as CD
from components import chart_data as CDT
from components import charts as C  # Pace-Chart shared with Standort
from components import (
    download_button,
    inject_brand_css,
    lazy_section,
    preload_all_button,
    render_notepad,
    render_toc,
    section,
    sync_snapshot_override,
)
from components import global_charts as GC
from components import global_tables as GT
from components.alerts import alert_card, alert_cards
from components.brand import hero
from components.export import register_section, reset_export
from components.tooltips import (
    KPI_GLOBAL_IST_OLD,
    KPI_GLOBAL_IST_STAY,
    KPI_GLOBAL_PLAN,
    KPI_GLOBAL_SALES,
    chart_help,
)
from revenueblindspots import helpers as H

st.set_page_config(
    page_title="Global Report",
    page_icon="🌍",
    layout="wide",
)
inject_brand_css()
CD.apply_stayery_style_once()
sync_snapshot_override()
CD.keep_session_state_alive()  # MUST run before any widget renders this page

PAGE = "global"
st.session_state["__page"] = PAGE

hero(
    eyebrow="Portfolio · Revenue-Recap",
    title="Global Report",
    subtitle="Standortübergreifend - IST vs. PLAN vs. Vorjahr, "
    "Pace-by-Month, Channel-Mix, Heatmaps.",
)


def _quarter_bounds(year: int, q: int) -> tuple[pd.Timestamp, pd.Timestamp]:
    start_month = (q - 1) * 3 + 1
    start = pd.Timestamp(year=year, month=start_month, day=1)
    end_month = start_month + 2
    end = pd.Timestamp(year=year, month=end_month, day=1) + pd.offsets.MonthEnd(0)
    return start, end


# ============================== Sidebar filter =============================
with st.sidebar:
    st.header("Filter")

    meta = CD.get_metadata()
    if not meta:
        st.error("Kein Snapshot - bitte erst `Daten aktualisieren` ausführen.")
        st.stop()

    all_props = meta.get("properties") or H.all_properties()

    _today = pd.Timestamp.today().normalize()
    _current_year = int(_today.year)
    _current_quarter = ((int(_today.month) - 1) // 3) + 1
    _previous_year = _current_year - 1
    _current_quarter_start, _current_quarter_end = _quarter_bounds(_current_year, _current_quarter)
    _previous_quarter_start, _previous_quarter_end = _quarter_bounds(
        _previous_year, _current_quarter
    )

    # `key=` ist nötig damit Filter bei Tab-Wechsel erhalten bleiben.
    props_pick = st.multiselect(
        "Standorte",
        options=all_props,
        default=all_props,
        key="global_props_pick",
    )
    mode = st.radio("Periode", ["Quartal", "Freie Periode"], horizontal=True, key="global_mode")

    with st.form("global_dates", clear_on_submit=False, border=False):
        if mode == "Quartal":
            c1, c2 = st.columns(2)
            with c1:
                y_new = st.number_input(
                    "Jahr (aktuell)",
                    min_value=2018,
                    max_value=2035,
                    value=_current_year,
                    step=1,
                    key="global_y_new",
                )
                q_new = st.selectbox(
                    "Quartal",
                    [1, 2, 3, 4],
                    index=_current_quarter - 1,
                    key="q_new",
                )
            with c2:
                y_old = st.number_input(
                    "Jahr (Vergleich)",
                    min_value=2018,
                    max_value=2035,
                    value=_previous_year,
                    step=1,
                    key="global_y_old",
                )
                q_old = st.selectbox(
                    "Quartal",
                    [1, 2, 3, 4],
                    index=_current_quarter - 1,
                    key="q_old",
                )
            start_new, end_new = _quarter_bounds(int(y_new), int(q_new))
            start_old, end_old = _quarter_bounds(int(y_old), int(q_old))
            period_tag_new = f"Q{q_new} {y_new}"
            period_tag_old = f"Q{q_old} {y_old}"
        else:
            c1, c2 = st.columns(2)
            with c1:
                st.caption("OLD")
                so = st.date_input(
                    "Start",
                    value=_previous_quarter_start.date(),
                    key="go_start",
                )
                eo = st.date_input(
                    "Ende",
                    value=_previous_quarter_end.date(),
                    key="go_end",
                )
            with c2:
                st.caption("NEW")
                sn = st.date_input(
                    "Start",
                    value=_current_quarter_start.date(),
                    key="gn_start",
                )
                en = st.date_input(
                    "Ende",
                    value=_current_quarter_end.date(),
                    key="gn_end",
                )
            start_old, end_old = pd.Timestamp(so), pd.Timestamp(eo)
            start_new, end_new = pd.Timestamp(sn), pd.Timestamp(en)
            period_tag_new = f"{start_new:%d.%m.%Y}–{end_new:%d.%m.%Y}"
            period_tag_old = f"{start_old:%d.%m.%Y}–{end_old:%d.%m.%Y}"

        green_threshold = st.slider(
            "Schwelle 🟢 (≥ PLAN + %)",
            min_value=0.5,
            max_value=10.0,
            value=2.0,
            step=0.5,
            key="global_green_pct",
        )
        red_threshold = st.slider(
            "Schwelle 🔴 (≤ PLAN − %)",
            min_value=-30.0,
            max_value=-1.0,
            value=-10.0,
            step=1.0,
            key="global_red_pct",
        )

        st.checkbox(
            "Späte Öffner einbeziehen",
            value=False,
            key="global_include_late_openers",
            help="Standorte deren Eröffnung nach dem OLD-Periodenende liegt "
            "werden standardmäßig ausgeschlossen - sonst verfälschen "
            "0-€-Zeilen die Totale & YoY-Vergleiche. Aktivieren wenn "
            "du sie für eine reine NEW-Sicht trotzdem sehen willst.",
        )

        st.checkbox(
            "Storno + No-Show einbeziehen",
            value=False,
            key="global_include_cancellations",
            help="Default: aus - alle KPIs/Tabellen sind realized-only "
            "Aktivieren → alle Buchungen zählen, "
            "auch später stornierte und nicht-erschienene. Hinweis: in der "
            "As-of-Sicht (§8) wirken beide point-in-time - ein Storno zählt bis "
            "zum cancel_time, ein No-Show bis zur Anreise (erst dort wird das "
            "Nicht-Erscheinen bekannt).",
        )

        st.form_submit_button("Recap aktualisieren", use_container_width=True)

    GT.GREEN_PCT = green_threshold
    GT.RED_PCT = red_threshold

    st.divider()
    st.caption("Sektionen 4-6 laden erst auf Klick.")
    preload_all_button(
        [5, 6, "7.A", "7.B", "7.C", "7.D"],
        label="Alle Sektionen laden",
    )
    CD.cache_clear_button()

    if not props_pick:
        st.warning("Bitte mindestens einen Standort wählen.")
        st.stop()

    st.divider()
    _plan_status = (
        "BigQuery-Snapshot (plan.parquet)" if CD.get_active_plan() else "– (kein Plan-Snapshot)"
    )
    st.caption(f"Plan: {_plan_status}")
    st.caption(f"Snapshot vom **{str(meta.get('refreshed_at', '?'))[:10]}**")

render_notepad(PAGE)


YEAR_OLD = start_old.year
YEAR_NEW = start_new.year
PERIOD_TAG = f"{period_tag_new} vs. {period_tag_old}"
plan_dict = CD.get_active_plan()  # Plan aus BigQuery-Snapshot (plan.parquet)
SNAP_TAG = CD.snapshot_tag()
SNAP_DATE = pd.Timestamp(str(meta.get("refreshed_at", ""))[:10] or pd.Timestamp.today().date())
props_tag = "+".join(sorted(props_pick)) if len(props_pick) < 11 else "all"


def _ck(section_id: str) -> str:
    # WICHTIG: Storno/No-Show-Toggle muss im Cache-Key stehen, sonst liefern die
    # gecachten Chart-PNGs beim Umschalten den alten (realized-only) Stand.
    _c = int(bool(st.session_state.get("global_include_cancellations", False)))
    return (
        f"glb::{SNAP_TAG}::{props_tag}::{start_old.date()}::{end_old.date()}"
        f"::{start_new.date()}::{end_new.date()}::c{_c}::{section_id}"
    )


@st.cache_data(ttl=3600, show_spinner=False, max_entries=4)
def _sc_export_xlsx(
    key: str,
    _nightly,
    _props,
    _sN,
    _eN,
    _sO,
    _eO,
    _csN,
    _ceN,
    _csO,
    _ceO,
    _asN,
    _asO,
    _incl,
) -> bytes:
    """Baut die §8-Export-Excel (Bytes). ``key`` steuert den Cache; die großen
    Argumente (``_nightly`` etc.) sind underscore-prefixed → werden nicht gehasht."""
    import io

    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    frames = GT.stay_created_export_frames(
        _nightly, _props, _sN, _eN, _sO, _eO, _csN, _ceN, _csO, _ceO, _asN, _asO, _incl
    )
    info = pd.DataFrame(
        {
            "Sheet / Flag": [
                "1 Stay NEW (alle)",
                "2 Stay OLD (alle)",
                "3 Stay+Created NEW",
                "4 Stay+Created OLD",
                "teilweise_im_stayfenster",
                "storno_vor/nach_stichtag",
                "noshow_vor/nach_stichtag",
                "zaehlt_in_8A",
                "revenue",
                "brutto_x1.07",
            ],
            "Bedeutung": [
                "Alle Nächte mit Aufenthalt im Stay-Fenster (aktuelles Jahr) - ohne Creation-/As-of-Filter.",
                "Wie 1, gespiegeltes Vorjahr (OLD).",
                "Wie 1, zusätzlich auf das Erstellungs-Fenster eingeschränkt - mit Point-in-Time-Flags.",
                "Wie 3, gespiegeltes Vorjahr (OLD).",
                "WAHR = die Buchung hat Nächte außerhalb des Stay-Fensters (nur ein Teil fällt rein).",
                "Storno (cancel_time) nach bzw. am/vor dem As-of-Stichtag.",
                "No-Show wird am arrival aufgelöst - nach bzw. am/vor dem Stichtag.",
                "WAHR = zählt in den §8-Tabellen (nach As-of + aktuellem Storno/No-Show-Toggle).",
                "Nacht-Netto (baseAmount_netAmount) - ohne Services/Extras.",
                "Brutto = Netto × 1,07 (nur Übernachtung; Frühstück/Parkplatz NICHT enthalten).",
            ],
        }
    )
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl", datetime_format="YYYY-MM-DD HH:MM") as xw:
        info.to_excel(xw, sheet_name="0 Info", index=False)
        for name, df in frames.items():
            out = df if len(df) else pd.DataFrame({"Hinweis": ["keine Daten in diesem Fenster"]})
            out.to_excel(xw, sheet_name=name, index=False)
        for ws in xw.book.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for i in range(1, ws.max_column + 1):
                ws.cell(row=1, column=i).font = Font(bold=True)
                ws.column_dimensions[get_column_letter(i)].width = 17
    return bio.getvalue()


# ============================== Data load ==================================
with st.spinner("Lade Daten aus dem Parquet-Snapshot …"):
    pull_start, pull_end = H.union_period((start_old, end_old), (start_new, end_new))
    pace_pull_start = pd.Timestamp(f"{YEAR_OLD}-01-01")
    pace_pull_end = pd.Timestamp(f"{YEAR_NEW}-12-31")
    pull_start = min(pull_start, pace_pull_start)
    pull_end = max(pull_end, pace_pull_end)
    nightly = CD.get_timeslices(start=pull_start, end=None, properties=props_pick)

reset_export(PAGE)


# Proactive: Standorte die in OLD-Periode noch nicht offen waren werden
# standardmäßig aus der Computation entfernt
_late_openers = H.properties_without_old_data(props_pick, end_old)
if _late_openers:
    _lines = "  · ".join(
        f"{pc} ({H.city(pc)}, eröffnet {H.opening_date(pc):%d.%m.%Y})" for pc in _late_openers
    )
    _include_late = st.session_state.get("global_include_late_openers", False)
    if _include_late:
        alert_card(
            f"{_lines}\n\nDiese Standorte sind in{period_tag_old} "
            f"noch nicht offen - Spalten zeigen 0 €. Toggle in der Sidebar "
            f'„Späte Öffner einbeziehen" zum Ausschluss.',
            kind="warning",
            title=f"Späte Öffner einbezogen ({period_tag_old})",
        )
    else:
        alert_card(
            f"{_lines}\n\nWurden für diese Analyse automatisch ausgeschlossen, "
            f"weil in {period_tag_old} noch nicht offen. Toggle in der "
            f'Sidebar „Späte Öffner einbeziehen" um sie trotzdem zu zeigen.',
            kind="info",
            title=f"Späte Öffner ausgeschlossen ({period_tag_old})",
        )
        props_pick = [p for p in props_pick if p not in _late_openers]
        if not props_pick:
            st.error(
                "Nach Ausschluss der späten Öffner ist die Standort-Auswahl "
                "leer. Bitte OLD-Periode später wählen oder „Späte Öffner "
                'einbeziehen" in der Sidebar aktivieren.'
            )
            st.stop()
        nightly = nightly[~nightly["property_code"].isin(_late_openers)]
        props_tag = "+".join(sorted(props_pick)) if len(props_pick) < 11 else "all"

# Warnung: NEW-Periode in der Zukunft.
if start_new > SNAP_DATE:
    alert_card(
        f"NEW-Periode ({period_tag_new}) liegt ganz in der Zukunft des "
        f"Snapshots ({SNAP_DATE:%d.%m.%Y}). Realisiertes Revenue = 0, nur "
        f"Forward-Bookings (Sales-Sicht) sind sichtbar.",
        kind="warning",
        title="NEW-Periode liegt in der Zukunft",
    )

# Non-fatale Warnung: gewählte Periode reicht vor den Datenbestand des Snapshots
# zurück (Lookback-Limit) - davor existieren keine Buchungen, Werte sind unvollständig.
_data_start = H.snapshot_data_start(meta)
if _data_start is not None:
    _before = [
        (lbl, s)
        for lbl, s in ((period_tag_old, start_old), (period_tag_new, start_new))
        if s < _data_start
    ]
    if _before:
        _txt = " · ".join(f"{lbl} beginnt {s:%d.%m.%Y}" for lbl, s in _before)
        alert_card(
            f"{_txt} - der Snapshot enthält aber erst Daten ab "
            f"{_data_start:%d.%m.%Y}. Der Zeitraum davor ist leer, die Werte (v.a. "
            f"Vorjahr) sind dadurch unvollständig bzw. zu niedrig. Für einen "
            f"vollständigen Rückblick den Snapshot mit größerem Lookback neu ziehen.",
            kind="warning",
            title="Periode reicht vor den verfügbaren Datenbestand zurück",
        )


# ============================== Build recap data ==========================
# Storno/No-Show-Toggle aus der Sidebar - bestimmt ob die Tabellen realized-only
# rechnen (default) oder alle Buchungen reinnehmen.
_include_cancellations = bool(st.session_state.get("global_include_cancellations", False))
_scope_caption = (
    "Scope: alle Buchungen (inkl. Storno + No-Show)"
    if _include_cancellations
    else "Scope: realized-only (Storno + No-Show ausgeschlossen)"
)
st.caption(_scope_caption)

disp_stay, raw_stay = GT.performance_by_stay(
    nightly,
    props_pick,
    start_new,
    end_new,
    start_old,
    end_old,
    YEAR_OLD,
    YEAR_NEW,
    period_tag_new,
    period_tag_old,
    plan=plan_dict,
    include_cancellations=_include_cancellations,
)
disp_chan_stay, raw_chan_stay = GT.channel_volume_by_stay(
    nightly,
    start_new,
    end_new,
    start_old,
    end_old,
    YEAR_OLD,
    YEAR_NEW,
    include_cancellations=_include_cancellations,
)
disp_created, raw_created = GT.performance_by_created(
    nightly,
    props_pick,
    start_new,
    end_new,
    start_old,
    end_old,
    YEAR_OLD,
    YEAR_NEW,
    include_cancellations=_include_cancellations,
)
disp_chan_created, raw_chan_created = GT.channel_volume_by_created(
    nightly,
    start_new,
    end_new,
    start_old,
    end_old,
    YEAR_OLD,
    YEAR_NEW,
    include_cancellations=_include_cancellations,
)


# ============================== TOC ========================================
_TOC = [
    (1, "Visual Scorecard"),
    (2, "Pace by Month"),
    ("3.A", "Performance Standorte (Erstellung)"),
    ("3.B", "Buchungskanäle (Erstellung)"),
    ("4.A", "Performance Standorte (Aufenthalt)"),
    ("4.B", "Buchungskanäle (Aufenthalt)"),
    (5, "IST vs PLAN · Pace-Fortschritt"),
    (6, "Channel-Mix Detail"),
    ("7.A", "Revenue-Heatmap Standort × Monat"),
    ("7.B", "Channel-Mix je Standort"),
    ("7.C", "Top-Movers"),
    ("7.D", "Channel × LOS (granular)"),
    (8, "Stay × Creation (As-of)"),
]
render_toc(_TOC)

st.caption(
    "**Datenbasis & Filter:** alle €-Werte = Stay-Date (`baseAmount_netAmount`). "
    "**§3** → **Erstellungsdatum** (created, Sales-Sicht). "
    "**§4 + Heatmaps/§7** → **Aufenthalt** (serviceDate, §4 mit PLAN). "
    "**Pace** = OTB-Rekonstruktion am Snapshot-Stichtag. Storno/No-Show + späte "
    "Öffner via Sidebar-Toggle."
)


# ============================== Executive Summary =========================
st.subheader("Executive Summary")


def _total(raw, col):
    if raw.empty:
        return 0.0
    sub = raw[raw["Standort"] == "Total"]
    return float(sub.iloc[0][col]) if len(sub) else 0.0


total_ist_stay = _total(raw_stay, "ist_new")
total_plan_stay = _total(raw_stay, "plan_new")
total_ly_stay = _total(raw_stay, "ist_old")
total_ist_cre = _total(raw_created, "ist_new")
total_ly_cre = _total(raw_created, "ist_old")

st.caption(
    "ℹ️ Die Kacheln folgen dem Sidebar-Toggle **Späte Öffner einbeziehen**: "
    "Standorte ohne Vorjahres-Daten sind nur enthalten, wenn der Toggle aktiv ist - "
    "dann steht ihr NEW-Umsatz aber keinem Vorjahr gegenüber, der YoY-Vergleich ist "
    "also leicht verzerrt. Toggle aus = späte Öffner überall raus (IST, PLAN, LY)."
)
c1, c2, c3, c4 = st.columns(4)
c1.metric(
    f"IST {period_tag_new} (Stay)",
    H.fmt_eur(total_ist_stay),
    delta=f"{(total_ist_stay - total_plan_stay):+,.0f} € vs PLAN".replace(",", "."),
    help=KPI_GLOBAL_IST_STAY,
)
c2.metric(f"PLAN {period_tag_new}", H.fmt_eur(total_plan_stay), help=KPI_GLOBAL_PLAN)
c3.metric(
    f"IST {period_tag_old} (Stay)",
    H.fmt_eur(total_ly_stay),
    delta=f"{(total_ist_stay - total_ly_stay):+,.0f} € YoY".replace(",", "."),
    help=KPI_GLOBAL_IST_OLD,
)
c4.metric(
    "Sales-Volumen NEW (Created)",
    H.fmt_eur(total_ist_cre),
    delta=f"{(total_ist_cre - total_ly_cre):+,.0f} € YoY".replace(",", "."),
    help=KPI_GLOBAL_SALES,
)

alerts = GT.auto_alerts(
    raw_stay,
    raw_created,
    YEAR_OLD,
    YEAR_NEW,
    include_cancellations=_include_cancellations,
)
st.markdown("**Automatische Alerts & Highlights**")
alert_cards(alerts)

_ALERT_ICONS = {"alert": "⚠", "warning": "▲", "success": "✓", "info": "ℹ"}
if alerts:
    _alert_lines = "**Alerts:**\n" + "\n".join(
        f"- {_ALERT_ICONS.get(h['kind'], '•')} **{h.get('title', '')}** - {h.get('message', '')}"
        for h in alerts
    )
else:
    _alert_lines = "_Keine automatischen Alerts._"

register_section(
    "exec_summary",
    "Executive Summary",
    body_markdown=(
        f"**IST {period_tag_new}:** {H.fmt_eur(total_ist_stay)}  ·  "
        f"**PLAN:** {H.fmt_eur(total_plan_stay)}  ·  "
        f"**IST {period_tag_old}:** {H.fmt_eur(total_ly_stay)}  \n"
        f"**Sales-Volumen NEW (Created):** {H.fmt_eur(total_ist_cre)}  "
        f"({total_ist_cre - total_ly_cre:+,.0f} € vs {period_tag_old})\n\n" + _alert_lines
    ),
    page=PAGE,
)
st.divider()


# ===== 1 · Visual Scorecard ================================================
with section(
    1,
    "Visual Scorecard",
    subtitle=f"IST vs PLAN je Standort · {PERIOD_TAG}",
    description=(
        "Bar-Länge = IST-Revenue, schwarzer Tick = PLAN, "
        "gestrichelte graue Linie = Vorjahr. Farb-Schwellen "
        "kommen aus den Slidern in der Sidebar. Grau = kein PLAN."
    ),
):
    png = CD.chart_png(
        _ck(f"scorecard::g{green_threshold}::r{red_threshold}"),
        GC.visual_scorecard,
        GT.with_code_labels(raw_stay),
        YEAR_OLD,
        YEAR_NEW,
        period_tag_new,
        green_threshold,
        red_threshold,
    )
    st.image(png, use_container_width=False)
    CD.data_table_expander(disp_stay, filename="global_scorecard")
    register_section(
        "scorecard", "1 · Visual Scorecard", chart_png=png, table_df=disp_stay, page=PAGE
    )

    chart_help("scorecard")
# ===== 2 · Pace by Month (NEU) ============================================
with section(
    2,
    "Pace by Month",
    subtitle=f"Alle gewählten Standorte · {YEAR_OLD} vs {YEAR_NEW} · Stand {SNAP_DATE:%d.%m.%Y}",
    description=(
        f"Pro Monat: **{YEAR_OLD}/EoM** = finale Vorjahres-Realität · "
        f"**{YEAR_OLD}/Today** = Stand `{SNAP_DATE:%d.%m.%Y}` im Vorjahr "
        f"(was war on-the-books) · **{YEAR_NEW}/Today** = aktueller Stand."
    ),
):
    pace_df = H.pace_by_month(nightly, YEAR_OLD, YEAR_NEW, SNAP_DATE, properties=props_pick)
    label = (
        "Alle Standorte" if len(props_pick) == len(all_props) else f"{len(props_pick)} Standorte"
    )
    png = CD.chart_png(_ck("pace"), C.pace_by_month_chart, pace_df, label, YEAR_OLD, YEAR_NEW)
    st.image(png, use_container_width=False)
    CD.data_table_expander(pace_df, filename=f"global_pace_{YEAR_NEW}")
    register_section("pace_month", "2 · Pace by Month", chart_png=png, table_df=pace_df, page=PAGE)

    chart_help("pace_month")
# ===== 3 · Nach Erstellungsdatum ==========================================
st.markdown("# 3 · Performance nach Erstellungsdatum")
st.caption("Sales-Sicht: alles was im Zeitraum gebucht wurde. ")


with section(
    "3.A",
    "Performance Standorte nach Erstellungsdatum",
    subtitle=f"IST {period_tag_new} vs IST {period_tag_old}",
):
    if disp_created.empty:
        alert_card("Keine Reservierungen im Erstellungs-Zeitraum.", kind="info")
    else:
        st.dataframe(disp_created, hide_index=True, use_container_width=True)
        if not raw_created.empty:
            no_total = raw_created[raw_created["Standort"] != "Total"]
            if len(no_total):
                worst = no_total.nsmallest(1, "d_pct")
                best = no_total.nlargest(1, "d_pct")
                txts = []
                if len(worst):
                    w = worst.iloc[0]
                    txts.append(f"🔴 **{w['Standort']}** {w['d_pct']:+.1f}% YoY")
                if len(best):
                    b = best.iloc[0]
                    txts.append(f"🟢 **{b['Standort']}** {b['d_pct']:+.1f}% YoY")
                if txts:
                    alert_card(
                        " · ".join(txts), kind="info", title="Sales-Bewegungen (nach Erstellung)"
                    )
    register_section(
        "perf_created", "3.A · Performance Standorte (Erstellung)", table_df=disp_created, page=PAGE
    )
    chart_help("perf_created")


with section("3.B", "Buchungskanäle nach Erstellungsdatum"):
    if disp_chan_created.empty:
        alert_card("Keine Channel-Daten nach Erstellungsdatum.", kind="info")
    else:
        st.dataframe(disp_chan_created, hide_index=True, use_container_width=True)
    register_section(
        "chan_created", "3.B · Buchungskanäle (Erstellung)", table_df=disp_chan_created, page=PAGE
    )
    chart_help("chan_created")


# ===== 4 · Nach Aufenthaltsdatum ==========================================
st.markdown("# 4 · Performance nach Aufenthaltsdatum")
st.caption("PLAN-Vergleich.")


with section(
    "4.A",
    "Performance Standorte nach Aufenthalt",
    subtitle=f"IST {period_tag_new} vs PLAN vs IST {period_tag_old}",
):
    if disp_stay.empty:
        alert_card("Keine Übernachtungen im Aufenthalts-Zeitraum.", kind="info")
    else:
        st.dataframe(disp_stay, hide_index=True, use_container_width=True)
    register_section(
        "perf_stay", "4.A · Performance Standorte (Aufenthalt)", table_df=disp_stay, page=PAGE
    )
    chart_help("perf_stay")


with section("4.B", "Buchungskanäle nach Aufenthaltsdatum"):
    if disp_chan_stay.empty:
        alert_card("Keine Channel-Daten nach Aufenthaltsdatum.", kind="info")
    else:
        st.dataframe(disp_chan_stay, hide_index=True, use_container_width=True)
    register_section(
        "chan_stay", "4.B · Buchungskanäle (Aufenthalt)", table_df=disp_chan_stay, page=PAGE
    )
    chart_help("chan_stay")


# ===== 5 · IST vs PLAN · Pace-Fortschritt =================================
if lazy_section(
    5,
    "IST vs PLAN · Pace-Fortschritt",
    subtitle="Was ist bisher reingekommen vs Plan, und wie weit ist die Periode schon durch.",
):
    pace_df = GC.build_pace_table(raw_stay, start_new, end_new)
    if pace_df.empty:
        alert_card("Keine Pace-Daten verfügbar.", kind="info")
    else:
        pace_disp = pace_df.copy()
        for c in ("IST (€)", "PLAN (€)"):
            pace_disp[c] = pace_disp[c].map(H.fmt_eur)
        pace_disp["IST / PLAN (%)"] = pace_disp["IST / PLAN (%)"].map(
            lambda v: f"{v:.0f} %" if pd.notna(v) else "-"
        )
        pace_disp["Fortschritt Zeit (%)"] = pace_disp["Fortschritt Zeit (%)"].map(
            lambda v: f"{v:.0f} %"
        )
        st.dataframe(pace_disp, hide_index=True, use_container_width=True)

        # Chart bekommt Hotel-Codes statt Stadt-Namen (kompakter)
        pace_df_chart = GC.build_pace_table(GT.with_code_labels(raw_stay), start_new, end_new)
        png = CD.chart_png(
            _ck("pace_plan"), GC.pace_to_plan_chart, pace_df_chart, YEAR_NEW, period_tag_new
        )
        st.image(png, use_container_width=False)
        register_section(
            "pace_plan",
            "5 · IST vs PLAN · Pace-Fortschritt",
            chart_png=png,
            table_df=pace_df,
            page=PAGE,
        )
        chart_help("pace_plan")
# ===== 6 · Channel-Mix Detail =============================================
if lazy_section(6, "Channel-Mix Detail", subtitle="Donut + horizontale Top-Bars (nach Aufenthalt)"):
    if not raw_chan_stay.empty:
        png_donut = CD.chart_png(
            _ck("ch_donut"), GC.channel_mix_donuts, raw_chan_stay, YEAR_OLD, YEAR_NEW
        )
        st.image(png_donut, use_container_width=False)
        png_bars = CD.chart_png(
            _ck("ch_bars"), GC.channel_mix_bars, raw_chan_stay, YEAR_OLD, YEAR_NEW
        )
        st.image(png_bars, use_container_width=False)
        CD.data_table_expander(disp_chan_stay, filename=f"global_channel_detail_{YEAR_NEW}")
        register_section(
            "channel_detail",
            "6 · Channel-Mix Detail",
            chart_png=png_donut,
            table_df=disp_chan_stay,
            page=PAGE,
        )

        chart_help("channel_detail")
# ===== 7 · Supporting Insights ============================================
st.markdown("# 7 · Supporting Insights")
st.caption("Heatmaps und Top-Movers.")


if lazy_section("7.A", "Revenue-Heatmap Standort × Monat"):
    png = CD.chart_png(
        _ck("heat_loc_month"),
        GC.location_revenue_heatmap,
        nightly,
        props_pick,
        pull_start,
        pull_end,
        realized_only=not _include_cancellations,
        title_suffix=(
            " (nach Aufenthalt, inkl. Storno+No-Show)"
            if _include_cancellations
            else " (nach Aufenthalt, realized)"
        ),
    )
    st.image(png, use_container_width=False)
    _tbl_locrev = CDT.location_revenue_table(
        nightly, pull_start, pull_end, realized_only=not _include_cancellations
    )
    CD.data_table_expander(_tbl_locrev, filename="global_loc_x_month")
    register_section(
        "heat_loc_month",
        "7.A · Revenue-Heatmap Standort × Monat",
        chart_png=png,
        table_df=_tbl_locrev,
        page=PAGE,
    )

    chart_help("heat_loc_month")
if lazy_section("7.B", "Channel-Mix je Standort"):
    png = CD.chart_png(
        _ck("heat_chan_loc"),
        GC.channel_x_location_heatmap,
        nightly,
        start_new,
        end_new,
        realized_only=not _include_cancellations,
    )
    st.image(png, use_container_width=False)
    _tbl_chloc = CDT.channel_x_location_table(
        nightly, start_new, end_new, realized_only=not _include_cancellations
    )
    CD.data_table_expander(_tbl_chloc, filename="global_channel_x_location")
    register_section(
        "heat_chan_loc",
        "7.B · Channel-Mix je Standort",
        chart_png=png,
        table_df=_tbl_chloc,
        page=PAGE,
    )

    chart_help("heat_chan_loc")
if lazy_section("7.C", "Top-Movers · Δ Revenue YoY"):
    png = CD.chart_png(
        _ck("top_movers"), GC.top_movers, GT.with_code_labels(raw_stay), YEAR_OLD, YEAR_NEW
    )
    st.image(png, use_container_width=False)
    # Top-Movers-Tabelle: raw_stay ohne Total, sortiert nach Δ Revenue.
    _tbl_mv = (
        raw_stay[raw_stay["Standort"] != "Total"][
            ["Standort", "ist_new", "ist_old", "d_ly_eur", "d_ly_pct"]
        ]
        .sort_values("d_ly_eur")
        .reset_index(drop=True)
    )
    _tbl_mv.columns = [
        "Standort",
        f"IST {YEAR_NEW} (€)",
        f"IST {YEAR_OLD} (€)",
        "Δ Revenue (€)",
        "Δ Revenue (%)",
    ]
    CD.data_table_expander(_tbl_mv, filename="global_top_movers")
    register_section("top_movers", "7.C · Top-Movers", chart_png=png, table_df=_tbl_mv, page=PAGE)
    chart_help("top_movers")


if lazy_section(
    "7.D",
    "Channel × LOS · granular",
    subtitle="Top-Channels (Booking.com, Expedia, HRS, IBE, …) nach Aufenthaltsdauer",
):
    png = CD.chart_png(
        _ck("heat_ch_los_granular"),
        GC.channel_los_heatmap_granular,
        nightly,
        start_old,
        end_old,
        start_new,
        end_new,
        YEAR_OLD,
        YEAR_NEW,
        realized_only=not _include_cancellations,
    )
    st.image(png, use_container_width=False)
    _tbl_chlosg = CDT.channel_los_granular_table(
        nightly,
        start_old,
        end_old,
        start_new,
        end_new,
        YEAR_OLD,
        YEAR_NEW,
        realized_only=not _include_cancellations,
    )
    CD.data_table_expander(_tbl_chlosg, filename="global_channel_los_granular")
    register_section(
        "heat_ch_los_granular",
        "7.D · Channel × LOS (granular)",
        chart_png=png,
        table_df=_tbl_chlosg,
        page=PAGE,
    )
    chart_help("heat_ch_los_granular")


# ===== 8 · Stay × Creation (As-of) ========================================
# Hauptbasis = Aufenthaltsdatum (Sidebar). Zusätzlicher Erstellungsdatum-Filter
# DIREKT ÜBER der Tabelle (nicht Sidebar), je Jahr gespiegelt. Storno/No-Show =
# derselbe Sidebar-Toggle, aber POINT-IN-TIME (As-of) wie der Pace-Chart.
YEAR_DELTA = YEAR_NEW - YEAR_OLD

# Default-Creation-Fenster: der Kalendermonat unmittelbar vor dem NEW-Stay-Start
# (z.B. Stay Juli → gebucht im Juni). Nur Vorbelegung - der User passt an.
_def_cre_end_new = (start_new - pd.Timedelta(days=1)).normalize()
_def_cre_start_new = _def_cre_end_new.replace(day=1)

with section(
    8,
    "Stay × Creation (As-of)",
    subtitle=f"Aufenthalt {period_tag_new} vs {period_tag_old}, zusätzlich nach "
    f"Erstellungsdatum gefiltert · {YEAR_NEW} vs {YEAR_OLD}",
    description=(
        "**Doppelt gefiltert:** Hauptmenge sind die Buchungen mit **Aufenthalt im "
        "Stay-Fenster** (Sidebar). Der Filter unten schränkt zusätzlich auf ein "
        "**Erstellungs-Fenster** ein. Das ist je Vergleichsjahr gespiegelt, damit beide "
        "Jahre vergleichbar sind. Der Sidebar-Toggle **Storno + No-Show** wirkt "
        "hier **point-in-time** (Stand Stichtag), nicht über den finalen Status. "
        "**Storno** zählt bis zum cancel_time, **No-Show** bis zur **Anreise** - "
        "erst am Anreisetag ist das Nicht-Erscheinen bekannt. Praktische Folge: "
        "liegt das Erstellungs-Fenster ganz vor dem Aufenthalt, "
        "zählen No-Shows per Default mit; überschneiden "
        "sich beide Fenster, fallen bereits bekannte No-Shows wie Stornos raus."
    ),
):
    with st.form("glb_sc_form", clear_on_submit=False, border=True):
        st.markdown("**Erstellungsdatum-Filter** (gilt für beide Jahre)")
        st.caption(
            " Gilt zusätzlich zum Stay-Fenster aus der "
            "Sidebar. ⚠️ Die Liniengrafik **8.D** ist ausschließlich für "
            "**Jahr-zu-Jahr-Vergleiche** gedacht."
        )
        _cc1, _cc2, _cc3, _cc4 = st.columns(4)
        with _cc1:
            _vm = st.selectbox(
                "von · Monat",
                list(range(1, 13)),
                index=int(_def_cre_start_new.month) - 1,
                format_func=lambda m: f"{m:02d} · " + H.MONTH_ABBR_DE[m - 1],
                key="global_sc_cre_vm",
            )
        with _cc2:
            _vd = st.number_input(
                "von · Tag",
                min_value=1,
                max_value=31,
                value=int(_def_cre_start_new.day),
                step=1,
                key="global_sc_cre_vd",
            )
        with _cc3:
            _bm = st.selectbox(
                "bis · Monat",
                list(range(1, 13)),
                index=int(_def_cre_end_new.month) - 1,
                format_func=lambda m: f"{m:02d} · " + H.MONTH_ABBR_DE[m - 1],
                key="global_sc_cre_bm",
            )
        with _cc4:
            _bd = st.number_input(
                "bis · Tag",
                min_value=1,
                max_value=31,
                value=int(_def_cre_end_new.day),
                step=1,
                key="global_sc_cre_bd",
            )
        st.caption(
            "ℹ️ **Storno/No-Show** steuerst du über den Sidebar-Toggle "
            '**„Storno + No-Show einbeziehen"** - wenn aus zählt die **As-of-Sicht**: '
            "Stornos *nach* dem Stichtag zählen noch mit; ein No-Show gilt erst ab "
            "der **Anreise** als aufgelöst (davor zählt er mit egal ob der Filter an oder aus ist). "
            "Der wirksame As-of-Stichtag ist min(Fensterende, Snapshot)."
            "Wenn der Filter an ist, werden auch Buchungen inkludiert die später storniert haben."
        )
        st.form_submit_button("Analyse aktualisieren", use_container_width=True)

    # Monat+Tag auf das jeweilige Jahr setzen: NEW = Stay-Jahr (Sidebar). Der Tag
    # wird defensiv auf die Monatslänge geclampt (z.B. 31 in einem 30-Tage-Monat
    # → 30). Das OLD-Fenster wird weiter unten über mirror_years gespiegelt.
    cre_start_new = H.clamp_day(YEAR_NEW, int(_vm), int(_vd))
    cre_end_new = H.clamp_day(YEAR_NEW, int(_bm), int(_bd))

    if cre_end_new < cre_start_new:
        alert_card(
            'Das Erstellungs-Fenster ist leer: „von" liegt nach „bis". '
            "Bitte Datumsgrenzen korrigieren.",
            kind="warning",
            title="Ungültiges Erstellungs-Fenster",
        )
    else:
        cre_start_old = H.mirror_years(cre_start_new, YEAR_DELTA)
        cre_end_old = H.mirror_years(cre_end_new, YEAR_DELTA)
        snap_old = H.mirror_years(SNAP_DATE, YEAR_DELTA)
        # Stichtag = min(Fensterende, Snapshot), je Jahr gespiegelt (PO-Entscheid).
        asof_new = min(cre_end_new, SNAP_DATE)
        asof_old = min(cre_end_old, snap_old)

        _sc_scope_txt = (
            "alle Buchungen (inkl. Storno + No-Show)"
            if _include_cancellations
            else "realized-only (Storno + No-Show point-in-time raus)"
        )
        st.caption(
            f"**Stay-Fenster:** {period_tag_new} vs {period_tag_old}  ·  "
            f"**Erstellungs-Fenster:** {cre_start_new:%d.%m.%Y}–{cre_end_new:%d.%m.%Y} "
            f"(NEW) ↔ {cre_start_old:%d.%m.%Y}–{cre_end_old:%d.%m.%Y} (OLD)  ·  "
            f"**As-of-Stichtag:** {asof_new:%d.%m.%Y} (NEW) · {asof_old:%d.%m.%Y} (OLD)  ·  "
            f"**Scope:** {_sc_scope_txt}."
        )

        disp_sc_loc, raw_sc_loc = GT.performance_by_stay_created(
            nightly,
            props_pick,
            start_new,
            end_new,
            start_old,
            end_old,
            cre_start_new,
            cre_end_new,
            cre_start_old,
            cre_end_old,
            asof_new,
            asof_old,
            YEAR_OLD,
            YEAR_NEW,
            include_cancellations=_include_cancellations,
        )
        disp_sc_chan, raw_sc_chan = GT.channel_volume_by_stay_created(
            nightly,
            start_new,
            end_new,
            start_old,
            end_old,
            cre_start_new,
            cre_end_new,
            cre_start_old,
            cre_end_old,
            asof_new,
            asof_old,
            YEAR_OLD,
            YEAR_NEW,
            include_cancellations=_include_cancellations,
        )
        disp_sc_seg, raw_sc_seg = GT.segment_volume_by_stay_created(
            nightly,
            start_new,
            end_new,
            start_old,
            end_old,
            cre_start_new,
            cre_end_new,
            cre_start_old,
            cre_end_old,
            asof_new,
            asof_old,
            YEAR_OLD,
            YEAR_NEW,
            include_cancellations=_include_cancellations,
        )

        if raw_sc_loc.empty:
            alert_card(
                "Keine Buchungen mit Aufenthalt im Stay-Fenster, die im gewählten "
                "Erstellungs-Fenster (bis zum Stichtag) angelegt wurden.",
                kind="info",
            )
        else:
            st.markdown("**8.A · nach Standort** (YoY, ohne PLAN)")
            st.dataframe(disp_sc_loc, hide_index=True, use_container_width=True)

            st.markdown("**8.B · nach Buchungskanal**")
            if disp_sc_chan.empty:
                alert_card("Keine Channel-Daten im gewählten Fenster.", kind="info")
            else:
                st.dataframe(disp_sc_chan, hide_index=True, use_container_width=True)

            st.markdown("**8.C · nach Stay-Segment** (kurz ≤6 / mittel 7-28 / lang 29+)")
            if disp_sc_seg.empty:
                alert_card("Keine Segment-Daten im gewählten Fenster.", kind="info")
            else:
                st.dataframe(disp_sc_seg, hide_index=True, use_container_width=True)

            # 8.D · Liniengrafik: Revenue je Erstellungs-Tag, NEW vs OLD.
            scope_new = GT.stay_created_scope(
                nightly,
                start_new,
                end_new,
                cre_start_new,
                cre_end_new,
                asof_new,
                _include_cancellations,
            )
            scope_old = GT.stay_created_scope(
                nightly,
                start_old,
                end_old,
                cre_start_old,
                cre_end_old,
                asof_old,
                _include_cancellations,
            )

            st.markdown(
                f"**8.D · Revenue je Erstellungs-Tag** (NEW vs OLD)  nur Buchungen "
                f"mit **Aufenthalt im Stay-Fenster** ({period_tag_new} vs {period_tag_old})"
            )
            st.caption(
                "⚠️ Diese Grafik richtet NEW und OLD am Tages-Offset des "
                "Erstellungs-Fensters aus und ist deshalb **nur für "
                "Jahr-zu-Jahr-Vergleiche** sinnvoll (gleicher Monat/Tag, ein Jahr "
                "versetzt). Für Vergleiche mit anderem Periodenzuschnitt ist sie "
                "nicht aussagekräftig."
            )

            _sc_ck_base = _ck(
                f"stay_created::{cre_start_new.date()}::{cre_end_new.date()}"
                f"::{asof_new.date()}::{asof_old.date()}::c{int(_include_cancellations)}"
            )

            # Liniengrafik als @st.fragment: der Channel-/OTA-Filter aktualisiert
            # NUR diesen Block - sofort und ohne Ganzseiten-Rerun (kein Hochspringen,
            # kein Sidebar-Knopf nötig). Die Tabellen 8.A-8.C bleiben unberührt.
            # Use-Case: Wirkung von Marketing-Maßnahmen je OTA über die Buchungstage.
            # Daten werden als Default-Argumente „eingefroren", damit der Fragment-
            # Rerun mit exakt den Werten des letzten Voll-Runs läuft.
            @st.fragment
            def _sc_line_fragment(
                scope_new=scope_new,
                scope_old=scope_old,
                cre_start_new=cre_start_new,
                cre_start_old=cre_start_old,
                ck_base=_sc_ck_base,
                year_old=YEAR_OLD,
                year_new=YEAR_NEW,
                label_dates=f"{cre_start_new:%d.%m.}–{cre_end_new:%d.%m.%Y}",
            ):
                ch_pool = pd.concat(
                    [scope_new["channel_combo"], scope_old["channel_combo"]],
                    ignore_index=True,
                ).dropna()
                ota_options = sorted({GT._channel_label(c) for c in ch_pool.unique()})
                ota_pick = st.multiselect(
                    "Channel-Filter – nur für die Liniengrafik (z.B. einzelne OTAs)",
                    options=ota_options,
                    default=[],
                    key="global_sc_ota_pick",
                    help="Aktualisiert NUR die Liniengrafik - sofort, ohne die Seite "
                    "neu zu laden. Die Tabellen 8.A–8.C bleiben unverändert. Leer = "
                    "alle Channels (Gesamt-Linie). Praktisch, um nach einer "
                    "Marketing-Maßnahme die Entwicklung einzelner OTAs zu vergleichen.",
                )
                if ota_pick:
                    m_new = scope_new["channel_combo"].map(GT._channel_label).isin(ota_pick)
                    m_old = scope_old["channel_combo"].map(GT._channel_label).isin(ota_pick)
                    ln_new, ln_old = scope_new[m_new], scope_old[m_old]
                    ota_tag = ", ".join(ota_pick)
                else:
                    ln_new, ln_old = scope_new, scope_old
                    ota_tag = "alle Channels"

                ldf = GT.daily_created_line_data(ln_new, ln_old, cre_start_new, cre_start_old)
                if ldf.empty:
                    alert_card(
                        "Keine Tages-Daten für die Liniengrafik (ggf. Channel-Filter "
                        "zu eng gewählt).",
                        kind="info",
                    )
                    st.session_state.pop("_sc_line_export", None)
                    st.session_state.pop("_sc_purpose_export", None)
                    st.session_state.pop("_sc_count_export", None)
                    return
                key = f"{ck_base}::ota={'+'.join(sorted(ota_pick)) if ota_pick else 'all'}"
                png = CD.chart_png(
                    key,
                    GC.stay_created_daily_chart,
                    ldf,
                    year_old,
                    year_new,
                    f"{label_dates} · {ota_tag}",
                )
                st.image(png, use_container_width=False)
                CD.data_table_expander(ldf, filename=f"global_stay_created_daily_{year_new}")
                # Für den Markdown-Export zwischenspeichern; die Registrierung
                # passiert im Haupt-Run außerhalb des Fragments (keine Duplikate).
                st.session_state["_sc_line_export"] = {"png": png, "table": ldf}

                # 8.E · Deep-Dive Channel: Composition Business vs Privat.
                # Reagiert auf DENSELBEN OTA-/Channel-Filter wie die Liniengrafik
                # (läuft auf ln_new/ln_old), zeigt die Reisezweck-Zusammensetzung
                # je Erstellungs-Tag - absolutes Revenue (Flächenhöhe) + Anteile
                # (Panel-Titel), OLD vs NEW deckungsgleich nebeneinander.
                st.markdown(
                    f"**8.E · Composition Business vs Privat je Erstellungs-Tag** "
                    f"(NEW vs OLD)  {ota_tag}"
                )
                alert_card(
                    "Der **Reisezweck** (Business/Privat) ist oft erst **nach "
                    "Check-in** verlässlich bekannt - je nach OTA ist das Feld ein "
                    "Pflichtfeld oder eben nicht. Buchungen **ohne** Reisezweck "
                    "(sowie alles Nicht-Business) zählen hier zu **Privat** und "
                    "können den Privat-Anteil überzeichnen - vor allem bei jungen "
                    "Buchungstagen am rechten Rand. Anteile daher mit Vorsicht "
                    "interpretieren.",
                    kind="info",
                    title="TravelPurpose oft erst nach Check-in bekannt",
                )
                adf = GT.purpose_daily_area_data(ln_new, ln_old, cre_start_new, cre_start_old)
                if adf.empty:
                    alert_card(
                        "Keine Tages-Daten für die Composition-Grafik (ggf. "
                        "Channel-Filter zu eng gewählt).",
                        kind="info",
                    )
                    st.session_state.pop("_sc_purpose_export", None)
                else:
                    png_purpose = CD.chart_png(
                        f"{key}::purpose",
                        GC.purpose_composition_area_chart,
                        adf,
                        year_old,
                        year_new,
                        f"{label_dates} · {ota_tag}",
                    )
                    st.image(png_purpose, use_container_width=False)
                    CD.data_table_expander(adf, filename=f"global_stay_created_purpose_{year_new}")
                    st.session_state["_sc_purpose_export"] = {
                        "png": png_purpose,
                        "table": adf,
                    }

                # 8.F · gleiche Aufteilung, aber nach ANZAHL Buchungen statt
                # Revenue (eindeutige Reservierungen). Gleicher OTA-/Channel-
                # Filter wie 8.D/8.E; beide Jahre als gruppierte Balken vergleichbar.
                st.markdown(f"**8.F · Anzahl Buchungen je Reisezweck** (NEW vs OLD) — {ota_tag}")
                cdf = GT.purpose_booking_counts(ln_new, ln_old)
                if cdf.empty:
                    alert_card(
                        "Keine Buchungen für die Count-Grafik (ggf. Channel-Filter "
                        "zu eng gewählt).",
                        kind="info",
                    )
                    st.session_state.pop("_sc_count_export", None)
                else:
                    png_count = CD.chart_png(
                        f"{key}::count",
                        GC.purpose_booking_count_chart,
                        cdf,
                        year_old,
                        year_new,
                        f"{label_dates} · {ota_tag}",
                    )
                    st.image(png_count, use_container_width=False)
                    CD.data_table_expander(cdf, filename=f"global_stay_created_counts_{year_new}")
                    st.session_state["_sc_count_export"] = {"png": png_count, "table": cdf}

            _sc_line_fragment()

            register_section(
                "stay_created_loc",
                "8.A · Stay × Creation (Standort)",
                table_df=disp_sc_loc,
                page=PAGE,
            )
            register_section(
                "stay_created_chan",
                "8.B · Stay × Creation (Channel)",
                table_df=disp_sc_chan,
                page=PAGE,
            )
            register_section(
                "stay_created_seg",
                "8.C · Stay × Creation (Stay-Segment)",
                table_df=disp_sc_seg,
                page=PAGE,
            )
            _sc_line_exp = st.session_state.get("_sc_line_export")
            if _sc_line_exp:
                register_section(
                    "stay_created_daily",
                    "8.D · Revenue je Erstellungs-Tag",
                    chart_png=_sc_line_exp["png"],
                    table_df=_sc_line_exp["table"],
                    page=PAGE,
                )
            _sc_purpose_exp = st.session_state.get("_sc_purpose_export")
            if _sc_purpose_exp:
                register_section(
                    "stay_created_purpose",
                    "8.E · Composition Business vs Privat",
                    chart_png=_sc_purpose_exp["png"],
                    table_df=_sc_purpose_exp["table"],
                    page=PAGE,
                )
            _sc_count_exp = st.session_state.get("_sc_count_export")
            if _sc_count_exp:
                register_section(
                    "stay_created_counts",
                    "8.F · Anzahl Buchungen je Reisezweck",
                    chart_png=_sc_count_exp["png"],
                    table_df=_sc_count_exp["table"],
                    page=PAGE,
                )

            # 8.G · Roh-Download aller Timeslices dieser §8-Sicht (zum Selber-Filtern).
            st.markdown(
                "**8.G · Download** alle Timeslices dieser Sicht als Excel "
                "(eine Zeile je Nacht). 4 Tabellenblätter + Info: Stay-Fenster (NEW/OLD) "
                "und zusätzlich Creation-gefiltert (NEW/OLD), inkl. Flags "
                "(teilweise im Stay-Fenster, Storno/No-Show vor/nach Stichtag, zählt in 8.A)."
            )
            _sc_xls_key = _ck(
                f"sc_export::{cre_start_new.date()}::{cre_end_new.date()}"
                f"::{asof_new.date()}::{asof_old.date()}"
            )
            _sc_xls = _sc_export_xlsx(
                _sc_xls_key,
                nightly,
                props_pick,
                start_new,
                end_new,
                start_old,
                end_old,
                cre_start_new,
                cre_end_new,
                cre_start_old,
                cre_end_old,
                asof_new,
                asof_old,
                _include_cancellations,
            )
            st.download_button(
                "Excel Download",
                data=_sc_xls,
                file_name=f"global_stay_created_{start_new:%Y%m%d}_vs_{start_old:%Y%m%d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    chart_help("stay_created")


# ===== Export ==============================================================
st.divider()
st.subheader("Bericht exportieren")
download_button(
    page_title=f"Global Report · {PERIOD_TAG}",
    highlights=alerts,
    filename=f"global_recap_{start_new:%Y%m%d}.md",
    page=PAGE,
)

CD.collect()
