"""Pickup / Vorlauf-Analyse - Stay × Creation Booking-Pace.

Kombiniert Aufenthalts- und Erstellungsdatum: Wie viel Umsatz eines Stay-Zeitraums
wurde in einem gewählten Erstellungs-Fenster gebucht - YoY, auf denselben Vorlauf
normiert. Funktioniert vorwärts (füllender Zukunftsmonat) wie rückwärts (einen
vergangenen Monat im Nachhinein auswerten). Löst den früheren §8-Block des
Global Reports ab.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(_REPO_ROOT / "src"))
sys.path.insert(0, str(_REPO_ROOT / "streamlit_app"))

import pandas as pd
import streamlit as st

from components import cached_data as CD
from components import global_charts as GC
from components import global_tables as GT
from components import inject_brand_css, sync_snapshot_override
from components.alerts import alert_card
from components.brand import hero
from components.global_tables import CancelMode
from revenueblindspots import helpers as H

st.set_page_config(page_title="Pickup / Vorlauf-Analyse", page_icon="📈", layout="wide")
inject_brand_css()
CD.apply_stayery_style_once()
sync_snapshot_override()
CD.keep_session_state_alive()

PAGE = "pickup"
st.session_state["__page"] = PAGE

hero(
    eyebrow="Portfolio · Booking-Pace",
    title="Pickup / Vorlauf-Analyse",
    subtitle="Stay × Creation: Wie viel Umsatz eines Aufenthalts-Zeitraums war "
    "zu einem Stichtag schon gebucht - im Jahresvergleich auf denselben Vorlauf "
    "normiert. Vorwärts (füllender Zukunftsmonat) wie rückwärts (vergangenen "
    "Monat auswerten).",
)

with st.expander("Was zeigt diese Seite? (kurz erklärt)", expanded=False):
    st.markdown(
        """
**Zwei Datumsachsen kombiniert.** *Stay-Datum* = wann der Gast übernachtet;
*Erstellungs-Datum* = wann gebucht wurde. Diese Seite beantwortet:
**Wie viel Umsatz eines Aufenthalts-Zeitraums war zu einem Stichtag schon
gebucht - und wie schlägt sich das gegen das Vorjahr?**

**Pickup-Anteil - die Kernzahl.**
`Pickup-Anteil = Erstellt-im-Fenster ÷ OTB gesamt`

- *Erstellt-im-Fenster (€)* = Umsatz für Stays im Stay-Fenster, der im gewählten
  **Erstellungs-Fenster** gebucht wurde.
- *OTB gesamt (€)* = gesamter Stay-Umsatz **on-the-books** zum Stichtag, ohne
  Erstellungs-Filter. Beide zum **selben Stichtag** gemessen.

**Fairer Jahresvergleich.** Der Vorjahres-Stichtag ist der aktuelle Stichtag
minus ganze Jahre. Beide Jahre werden so auf **demselben Buchungsvorlauf**
verglichen - egal ob der Monat noch **in der Zukunft** liegt (füllt sich noch)
oder schon **vorbei** ist (im Nachhinein lernen, z.B. im Februar den Januar).

**Storno-Modus.** *All in* = alle Buchungen (inkl. später
storniert / No-Show). *All out* = nur realisierte (finaler Status). *As-of* =
point-in-time zum Stichtag (Storno bis `cancel_time`, No-Show bis Anreise). Für
vergangene Monate sind As-of und All out gleich; für Zukunftsmonate zeigt As-of
den Stand, wie er am Stichtag aussah.

**Creation-Fenster.** *Festes Fenster* = z.B. nur im Juni gebucht (eine konkrete
Vorlauf-Phase). *Alles bis Stichtag* = kompletter Vorlauf ohne untere Grenze.
"""
    )

_CMODE_LABELS = {
    "As-of (Stichtag, point-in-time)": CancelMode.AS_OF,
    "All out (realized-only)": CancelMode.ALL_OUT,
    "All in (inkl. Storno/No-Show)": CancelMode.ALL_IN,
}


# ============================== Sidebar filter =============================
with st.sidebar:
    st.header("Filter")

    meta = CD.get_metadata()
    if not meta:
        st.error("Kein Snapshot - bitte erst `Daten aktualisieren` ausführen.")
        st.stop()

    all_props = meta.get("properties") or H.all_properties()
    SNAP_DATE = pd.Timestamp(
        str(meta.get("refreshed_at", ""))[:10] or pd.Timestamp.today().date()
    )
    _today = pd.Timestamp.today().normalize()

    # Defaults: Fokus Zukunft → Stay = nächster Monat, Creation = Vormonat.
    _stay_first = _today.replace(day=1) + pd.offsets.MonthBegin(1)
    _stay_last = _stay_first + pd.offsets.MonthEnd(0)
    _cre_first = _stay_first - pd.offsets.MonthBegin(1)
    _cre_last = _cre_first + pd.offsets.MonthEnd(0)

    props_pick = st.multiselect(
        "Standorte", options=all_props, default=all_props, key="pu_props"
    )

    with st.form("pu_filters", clear_on_submit=False, border=True):
        st.markdown("**Stay-Fenster (aktuelles Jahr)**")
        _c1, _c2 = st.columns(2)
        with _c1:
            sn = st.date_input("Start", value=_stay_first.date(), key="pu_sn")
        with _c2:
            en = st.date_input("Ende", value=_stay_last.date(), key="pu_en")
        vgl_year = st.number_input(
            "Vergleichsjahr (OLD)",
            min_value=2018,
            max_value=2035,
            value=int(_stay_first.year) - 1,
            step=1,
            key="pu_vgl_year",
            help="Das Stay-Fenster wird um diese Jahresdifferenz gespiegelt "
            "(gleicher Monat/Tag, ein oder mehrere Jahre versetzt).",
        )

        st.markdown("**Erstellungs-Fenster**")
        cre_mode = st.radio(
            "Modus",
            ["Festes Fenster (von–bis)", "Alles bis Stichtag"],
            key="pu_cre_mode",
            help="Festes Fenster = z.B. nur im Juni gebucht. Alles bis Stichtag = "
            "kein unteres Datum, alle Buchungen bis zum Stichtag.",
        )
        _c3, _c4 = st.columns(2)
        with _c3:
            cv = st.date_input(
                "Creation von", value=_cre_first.date(), key="pu_cv"
            )
        with _c4:
            cb = st.date_input("Creation bis", value=_cre_last.date(), key="pu_cb")
        st.caption("Creation von/bis gelten nur im Modus Festes Fenster.")

        st.markdown("**Stichtag (As-of)**")
        asof_in = st.date_input(
            "Stichtag",
            value=SNAP_DATE.date(),
            key="pu_asof",
            help="Point-in-time-Grenze. Default = Snapshot. Für Rückwärts-Analysen "
            "frei wählbar. Der Vorjahres-Stichtag wird automatisch gespiegelt. "
            "Konvention: Der Stichtag ist ein GANZER Kalendertag (keine Uhrzeit) - "
            "Buchungen, die am Stichtag erstellt wurden, zählen mit; ein Storno "
            "mit Zeitstempel am Stichtag gilt als bereits bekannt (zählt raus).",
        )

        st.markdown("**Storno-Modus**")
        cmode_label = st.radio(
            "Storno / No-Show",
            list(_CMODE_LABELS.keys()),
            key="pu_cmode",
            help="As-of = point-in-time (Storno bis cancel_time, No-Show bis Anreise). "
            "All out = nur realisierte Buchungen (finaler Status). "
            "All in = alle Buchungen inkl. später storniert / No-Show.",
        )

        st.form_submit_button("Analyse aktualisieren", use_container_width=True)

    CD.cache_clear_button()
    st.divider()
    # Farbige Freshness-Ampel statt Text-Caption: gruen <5h, gelb 5-15h, rot >15h.
    CD.freshness_badge()

if not props_pick:
    st.warning("Bitte mindestens einen Standort wählen.")
    st.stop()

# ============================== Ableitungen ================================
start_new = pd.Timestamp(sn)
end_new = pd.Timestamp(en)
if end_new < start_new:
    alert_card("Stay-Fenster ist leer: Start liegt nach Ende.", kind="warning")
    st.stop()

YEAR_NEW = int(start_new.year)
YEAR_OLD = int(vgl_year)
YEAR_DELTA = YEAR_NEW - YEAR_OLD
if YEAR_DELTA == 0:
    alert_card("Vergleichsjahr muss sich vom Stay-Jahr unterscheiden.", kind="warning")
    st.stop()

start_old = H.mirror_years(start_new, YEAR_DELTA)
end_old = H.mirror_years(end_new, YEAR_DELTA)

asof_new = pd.Timestamp(asof_in)
asof_old = H.mirror_years(asof_new, YEAR_DELTA)

CMODE = _CMODE_LABELS[cmode_label]

if cre_mode == "Festes Fenster (von–bis)":
    cre_start_new = pd.Timestamp(cv)
    cre_end_new = pd.Timestamp(cb)
    cre_tag = f"{cre_start_new:%d.%m.%Y}–{cre_end_new:%d.%m.%Y}"
else:
    # Kein unteres Datum: 3 Jahre vor Stay-Start als praktische Untergrenze.
    cre_start_new = pd.Timestamp(year=YEAR_NEW - 3, month=1, day=1)
    cre_end_new = asof_new
    cre_tag = f"alles bis {asof_new:%d.%m.%Y}"

if cre_end_new < cre_start_new:
    alert_card("Erstellungs-Fenster ist leer: von liegt nach bis.", kind="warning")
    st.stop()

cre_start_old = H.mirror_years(cre_start_new, YEAR_DELTA)
cre_end_old = H.mirror_years(cre_end_new, YEAR_DELTA)

period_tag_new = f"{start_new:%d.%m.%Y}–{end_new:%d.%m.%Y}"
period_tag_old = f"{start_old:%d.%m.%Y}–{end_old:%d.%m.%Y}"

st.caption(
    f"**Stay-Fenster:** {period_tag_new} (NEW) ↔ {period_tag_old} (OLD)  ·  "
    f"**Erstellungs-Fenster:** {cre_tag}  ·  "
    f"**Stichtag:** {asof_new:%d.%m.%Y} (NEW) · {asof_old:%d.%m.%Y} (OLD)  ·  "
    f"**Storno-Modus:** {cmode_label}."
)
st.caption(
    "**Pickup-Anteil** = Erstellt-im-Fenster ÷ OTB gesamt (voller Stay-Umsatz, "
    "ohne Erstellungs-Filter) - beide zum selben Stichtag, daher ≤ 100 %. "
    "Zähler und Nenner sind YoY auf denselben Vorlauf gespiegelt."
)

# ============================== Daten laden ================================
with st.spinner("Lade Daten aus dem Parquet-Snapshot …"):
    _pull_start = min(start_old, start_new, cre_start_old, cre_start_new)
    nightly = CD.get_timeslices(start=_pull_start, end=None, properties=props_pick)
    plan_dict = CD.get_active_plan()  # PLAN aus BigQuery-Snapshot (plan.parquet)

if nightly is None or nightly.empty:
    alert_card("Keine Timeslices im gewählten Bereich.", kind="info")
    st.stop()

# ============================== Tabellen ===================================
_common = dict(
    cancel_mode=CMODE,
    pickup=True,
)
disp_loc, raw_loc = GT.performance_by_stay_created(
    nightly, props_pick, start_new, end_new, start_old, end_old,
    cre_start_new, cre_end_new, cre_start_old, cre_end_old,
    asof_new, asof_old, YEAR_OLD, YEAR_NEW, **_common,
)
disp_ch, raw_ch = GT.channel_volume_by_stay_created(
    nightly, start_new, end_new, start_old, end_old,
    cre_start_new, cre_end_new, cre_start_old, cre_end_old,
    asof_new, asof_old, YEAR_OLD, YEAR_NEW, **_common,
)
disp_seg, raw_seg = GT.segment_volume_by_stay_created(
    nightly, start_new, end_new, start_old, end_old,
    cre_start_new, cre_end_new, cre_start_old, cre_end_old,
    asof_new, asof_old, YEAR_OLD, YEAR_NEW, **_common,
)

if raw_loc.empty:
    alert_card(
        "Keine Buchungen mit Aufenthalt im Stay-Fenster im gewählten "
        "Erstellungs-Fenster / Storno-Modus.",
        kind="info",
    )
    st.stop()

_tot = raw_loc[raw_loc["property_code"] == "TOTAL"].iloc[0]
_otb_new = float(_tot["stay_new"])
_otb_old = float(_tot["stay_old"])
_erst_new = float(_tot["ist_new"])
_erst_old = float(_tot["ist_old"])
# Fenster-Pickup: NUR im Erstellungs-Fenster gebuchter Umsatz ÷ OTB.
_pu_new = (_erst_new / _otb_new * 100.0) if _otb_new > 0 else float("nan")
_pu_old = (_erst_old / _otb_old * 100.0) if _otb_old > 0 else float("nan")
_pu_delta = _pu_new - _pu_old if pd.notna(_pu_new) and pd.notna(_pu_old) else float("nan")

# Kumulierter Pickup: ALLES bis zum Ende des Erstellungs-Fensters gebuchte
# (keine untere Grenze) ÷ OTB - "wie viel des OTB stand Ende Juni schon?".
# Liegt das Fenster-Ende am/nach dem Stichtag, ist der Wert per Definition 100 %.
_CUM_FLOOR = pd.Timestamp("2000-01-01")
_cum_new = float(GT.stay_created_scope(
    nightly, start_new, end_new, _CUM_FLOOR, cre_end_new, asof_new, cancel_mode=CMODE
)["revenue"].sum())
_cum_old = float(GT.stay_created_scope(
    nightly, start_old, end_old, _CUM_FLOOR, cre_end_old, asof_old, cancel_mode=CMODE
)["revenue"].sum())
_pu_cum_new = (_cum_new / _otb_new * 100.0) if _otb_new > 0 else float("nan")
_pu_cum_old = (_cum_old / _otb_old * 100.0) if _otb_old > 0 else float("nan")
_pu_cum_delta = (
    _pu_cum_new - _pu_cum_old
    if pd.notna(_pu_cum_new) and pd.notna(_pu_cum_old)
    else float("nan")
)


def _pct(v: float) -> str:
    return f"{v:.1f} %" if pd.notna(v) else "–"


# OTB "heute" (= am Stichtag) mit REINER As-of-Logik - unabhängig vom
# Storno-Modus-Schalter, identisch zur Monats-Tabelle & zum Tages-Verlauf in §5:
# No-Shows raus, Storno löst zu cancel_time auf (Storno AM Stichtag = raus,
# Buchung erstellt AM Stichtag = rein; Stichtag = ganzer Kalendertag).
def _asof_window_otb(df: pd.DataFrame, s_: pd.Timestamp, e_: pd.Timestamp,
                     asof_: pd.Timestamp) -> float:
    sub = H.filter_period(df, s_, e_, "stay_date")
    if "is_no_show" in sub.columns:
        sub = sub[~sub["is_no_show"].astype(bool)]
    if sub.empty:
        return 0.0
    on = H.asof_on_the_books_mask(sub, asof_, include_cancellations=False)
    return float(sub.loc[on, "revenue"].sum())


_aotb_new = _asof_window_otb(nightly, start_new, end_new, asof_new)
_aotb_old = _asof_window_otb(nightly, start_old, end_old, asof_old)
_aotb_delta_abs = _aotb_new - _aotb_old
_aotb_delta_pct = (_aotb_new / _aotb_old - 1) * 100.0 if _aotb_old > 0 else float("nan")


# ============================== KPIs =======================================
st.markdown("## 1 · Headline-Kennzahlen")
st.markdown("*Portfolio-Summe über alle gewählten Standorte.*")
st.caption(
    "**Pickup kumuliert** = Anteil des OTB, der **bis zum Ende des "
    "Erstellungs-Fensters** gebucht war (ohne untere Grenze) - wie viel stand "
    "z.B. Ende Juni schon in den Büchern? **In Klammern: Fenster-Pickup** = nur der "
    "**im Erstellungs-Fenster** (z.B. 01.–30.06.) erstellte Umsatz ÷ OTB. "
    "**OTB Δ heute** = OTB des Stay-Fensters am Stichtag vs. Vorjahres-Stichtag "
    "(reine As-of-Logik: No-Shows raus, Storno bis `cancel_time` - unabhängig vom "
    "Storno-Modus-Schalter; identisch zu Tabelle + Tages-Verlauf in §5). "
    "**OTB gesamt** = zum Stichtag gebuchter Stay-Umsatz (NEW), unabhängig vom "
    "Erstellungs-Fenster. Endet das Fenster am/nach dem Stichtag, ist kumuliert = 100 %. "
    "Die Δ-Pickup-Werte (pp) stehen weiterhin in den Tabellen (§4)."
)
k1, k2, k3, k4 = st.columns(4)
k1.metric(
    f"Pickup kumuliert {YEAR_NEW}",
    _pct(_pu_cum_new),
    delta=f"(im Fenster: {_pct(_pu_new)})",
    delta_color="off",
)
k2.metric(
    f"Pickup kumuliert {YEAR_OLD}",
    _pct(_pu_cum_old),
    delta=f"(im Fenster: {_pct(_pu_old)})",
    delta_color="off",
)
k3.metric(
    f"OTB Δ heute vs {YEAR_OLD}",
    f"{_aotb_delta_pct:+.1f} %" if pd.notna(_aotb_delta_pct) else "–",
    delta=f"{_aotb_delta_abs:+,.0f} €".replace(",", "."),
)
k4.metric(f"OTB gesamt {YEAR_NEW}", H.fmt_eur(_otb_new))


# ============================== Buchungskurve ==============================
# Creation-day-Scopes + Linien-Daten immer berechnen (Chart + CSV-Download).
_scope_new = GT.stay_created_scope(
    nightly, start_new, end_new, cre_start_new, cre_end_new, asof_new, cancel_mode=CMODE
)
_scope_old = GT.stay_created_scope(
    nightly, start_old, end_old, cre_start_old, cre_end_old, asof_old, cancel_mode=CMODE
)
_line_df = GT.daily_created_line_data(_scope_new, _scope_old, cre_start_new, cre_start_old)

st.markdown("## 2 · Buchungskurve")
st.markdown("*Wie sich der gebuchte Anteil über die Zeit aufbaut - NEW vs. OLD.*")
_curve_axis = st.radio(
    "X-Achse",
    ["Erstellungs-Tag", "Tage vor Anreise (Lead-Time)"],
    horizontal=True,
    key="pu_curve_axis",
    help="Erstellungs-Tag = Kalendertag im Erstellungs-Fenster (an der Fenster-"
    "Ausrichtung gespiegelt). Tage vor Anreise = Lead-Time-Sicht: über Monate "
    "vergleichbar, unabhängig von Kalendertag und Monatslänge.",
)
if _curve_axis == "Erstellungs-Tag":
    st.caption(
        "Jeder Punkt = kumulierter Anteil des Stay-OTB, der **bis zu diesem "
        "Erstellungs-Tag** gebucht war. Liegt die NEW-Linie **über** OLD, buchen "
        "wir früher/stärker vor als im Vorjahr; darunter = wir hinken hinterher. "
        "Beide Jahre sind am Tages-Offset des Erstellungs-Fensters ausgerichtet. "
        "Am aussagekräftigsten im Modus Festes Fenster."
    )
    _curve = GT.pickup_pace_curve(_line_df, _otb_new, _otb_old, YEAR_NEW, YEAR_OLD)
    if _curve.empty:
        alert_card("Keine Tages-Daten für die Buchungskurve.", kind="info")
    else:
        st.line_chart(_curve, y_label="Kumulierter Pickup-Anteil (%)")
else:
    st.caption(
        "Anteil des Stay-OTB, der **mindestens so viele Tage vor Anreise** "
        "gebucht war (Lead-Time). **NEW über OLD = wir buchen weiter im Voraus** "
        "(mehr Umsatz früh gesichert). Über Monate hinweg vergleichbar, weil an "
        "der Anreise statt am Kalendertag ausgerichtet."
    )
    _otb_scope_new = GT.stay_only_scope(
        nightly, start_new, end_new, asof_new, cancel_mode=CMODE
    )
    _otb_scope_old = GT.stay_only_scope(
        nightly, start_old, end_old, asof_old, cancel_mode=CMODE
    )
    _lead = GT.pickup_leadtime_curve(
        _otb_scope_new, _otb_scope_old, _otb_new, _otb_old, YEAR_NEW, YEAR_OLD
    )
    if _lead.dropna(how="all").empty:
        alert_card("Keine Lead-Time-Daten im gewählten Fenster.", kind="info")
    else:
        st.line_chart(
            _lead, x_label="Tage vor Anreise", y_label="kum. Anteil des OTB (%)"
        )


# ============================== Pickup-Balken ==============================
st.markdown("## 3 · Pickup-Anteil je Kategorie")
st.markdown("*Wer liegt im Vorlauf vorn, wer hinkt hinterher?*")
st.caption(
    "Pickup-Anteil je Standort / Buchungskanal / Stay-Segment (ohne Total), "
    "NEW vs. OLD. Höhere Balken = mehr des Stay-Umsatzes bereits gesichert. "
    "Große NEW-über-OLD-Lücken = Standorte/Kanäle mit stärkerem Vorlauf als "
    "im Vorjahr; NEW unter OLD = Nachhol-Bedarf."
)
_cat = st.radio(
    "Kategorie",
    ["Standort", "Buchungskanal", "Stay-Segment"],
    horizontal=True,
    key="pu_bar_cat",
)
if _cat == "Standort":
    _bars = GT.pickup_bars_data(raw_loc, "Standort", "ist_new", "ist_old", YEAR_NEW, YEAR_OLD)
elif _cat == "Buchungskanal":
    _bars = GT.pickup_bars_data(raw_ch, "Channel", "rev_new", "rev_old", YEAR_NEW, YEAR_OLD)
else:
    _bars = GT.pickup_bars_data(raw_seg, "Segment", "rev_new", "rev_old", YEAR_NEW, YEAR_OLD)
if _bars.empty:
    alert_card("Keine Kategorie-Daten für die Balken.", kind="info")
else:
    st.bar_chart(_bars, stack=False, y_label="Pickup-Anteil (%)")


# ============================== Tabellen ===================================
st.markdown("## 4 · Tabellen")
st.markdown("*Dieselbe Struktur je Kategorie - Detail hinter den Charts.*")
with st.expander("Spalten einfach erklärt (Klartext)", expanded=False):
    st.markdown(
        f"""
Jede Zeile ist **eine Gruppe** (ein Standort, ein Buchungskanal oder ein
Stay-Segment):

- **Erstellt {YEAR_NEW} (€)** - Umsatz für diesen Zeitraum, der **im ausgewählten
  Buchungs-Fenster** reingekommen ist. Beispiel: Umsatz für
  Juli-Übernachtungen, der im Juni gebucht wurde.
- **Erstellt {YEAR_OLD} (€)** - genau dasselbe, aber im **Vorjahr** (zum
  Vergleichen).
- **Δ absolut (€)** - der Unterschied in Euro: dieses Jahr **minus** letztes
  Jahr. Plus = mehr als letztes Jahr, Minus = weniger.
- **Δ relativ (%)** - derselbe Unterschied in **Prozent**. Die Ampel zeigt es
  auf einen Blick: 🟢 deutlich besser, 🟠 ähnlich, 🔴 deutlich schlechter.
- **OTB {YEAR_NEW} (€)** - der **gesamte** zum Stichtag gebuchte Umsatz für den
  Zeitraum, **egal wann** gebucht wurde. „On-the-books" = steht schon fest
  auf den Büchern.
- **OTB {YEAR_OLD} (€)** - dasselbe fürs **Vorjahr**.
- **Pickup-Anteil {YEAR_NEW} (%)** - wie viel **Prozent** des gesamten
  Umsatzes **genau im gewählten Buchungs-Fenster** reinkam - **NICHT
  kumuliert**. 40 % heißt: 40 % des OTB wurde zwischen dem 01.06. und
  30.06. gebucht; was davor schon gebucht war, steckt hier nicht drin.
  Die kumulierte Sicht („wie viel stand Ende Juni insgesamt schon in den
  Büchern?") zeigen die Kacheln oben unter **Pickup kumuliert**.
- **Pickup-Anteil {YEAR_OLD} (%)** - dasselbe fürs Vorjahr: sind wir dieses
  Jahr **früher oder später** dran als damals?
- **Δ Pickup (pp)** - um wie viele **Prozentpunkte** unser Vorlauf besser (+)
  oder schlechter (−) ist als letztes Jahr. Plus = wir buchen früher.
- **Δ Anteil (pp)** *(nur Kanal/Segment)* - wie sich der **Anteil dieser
  Gruppe am Gesamt-Umsatz** verschoben hat (der „Kuchen-Anteil"). Das ist der
  **Mix**, nicht der Vorlauf.

*„pp" = Prozentpunkte: der einfache Abstand zweier Prozentzahlen. Von 60 % auf
66 % sind es +6 pp.*
"""
    )
st.markdown(f"**Nach Standort** · {YEAR_NEW} vs {YEAR_OLD}")
st.dataframe(disp_loc, hide_index=True, use_container_width=True)

st.markdown("**Nach Buchungskanal**")
if disp_ch.empty:
    alert_card("Keine Channel-Daten im gewählten Fenster.", kind="info")
else:
    st.dataframe(disp_ch, hide_index=True, use_container_width=True)

st.markdown("**Nach Stay-Segment** (kurz ≤6 / mittel 7-28 / lang 29+)")
if disp_seg.empty:
    alert_card("Keine Segment-Daten im gewählten Fenster.", kind="info")
else:
    st.dataframe(disp_seg, hide_index=True, use_container_width=True)


# ============================== Pace-to-PLAN ===============================
st.markdown("## 5 · Pace-to-PLAN · OTB vs. Ziel")
st.markdown("*Wie viel des Stay-PLANs ist zum Stichtag on-the-books?*")
st.caption(
    "IST = **OTB gesamt** (aktueller Stay-Umsatz zum Stichtag, gem. Storno-Modus) "
    "je Standort gegen den **PLAN** des Stay-Fensters. **Fortschritt Zeit** = wie "
    "viel der Periode am Stichtag verstrichen ist - liegt IST/PLAN darüber, sind "
    "wir dem Zeitplan **voraus**, darunter **hinterher**. Unabhängig vom "
    "Erstellungs-Fenster. (Früher Global Report §5.)"
)
if not plan_dict:
    alert_card(
        "Kein PLAN-Snapshot geladen - Pace-to-PLAN nicht verfügbar.", kind="info"
    )
else:
    _pace_src = raw_loc[raw_loc["property_code"] != "TOTAL"][
        ["Standort", "property_code", "stay_new"]
    ].copy()
    _pace_src["ist_new"] = _pace_src["stay_new"]
    _pace_src["plan_new"] = [
        H.plan_revenue(pc, start_new, end_new, plan=plan_dict)
        for pc in _pace_src["property_code"]
    ]
    _pace_df = GC.build_pace_table(_pace_src, start_new, end_new, today=asof_new)
    if _pace_df.empty:
        alert_card(
            "Keine Pace-Daten (kein PLAN für dieses Stay-Fenster).", kind="info"
        )
    else:
        _pace_disp = _pace_df.copy()
        for _c in ("IST (€)", "PLAN (€)"):
            _pace_disp[_c] = _pace_disp[_c].map(H.fmt_eur)
        for _c in ("IST / PLAN (%)", "Fortschritt Zeit (%)"):
            _pace_disp[_c] = _pace_disp[_c].map(
                lambda v: f"{v:.1f}" if pd.notna(v) else "-"
            )
        st.dataframe(_pace_disp, hide_index=True, use_container_width=True)
        _pace_key = (
            f"pu_pace::{start_new.date()}::{asof_new.date()}::{CMODE.value}"
            f"::{'+'.join(sorted(props_pick))}"
        )
        _pace_png = CD.chart_png(
            _pace_key, GC.pace_to_plan_chart, _pace_df, YEAR_NEW, period_tag_new
        )
        st.image(_pace_png, use_container_width=False)


# ============================== 5b · OTB je Stay-Monat (As-of) =============
st.markdown("### OTB je Stay-Monat · Stichtags-Vergleich")
st.caption(
    f"On-the-books je **Stay-Monat {YEAR_NEW}** am Stichtag **{asof_new:%d.%m.%Y}** "
    f"vs. Vorjahres-Monat am gespiegelten Stichtag **{asof_old:%d.%m.%Y}** - "
    "reine **As-of-Logik** (No-Shows raus, Storno löst zu `cancel_time` auf), "
    "unabhängig vom Storno-Modus-Schalter und vom Erstellungs-Fenster. "
    "**Stichtag-Konvention:** ganzer Kalendertag ohne Uhrzeit - am Stichtag "
    "erstellte Buchungen zählen **mit**, ein Storno mit Zeitstempel am Stichtag "
    "gilt als bereits bekannt (zählt **raus**). "
    "**Zeile eines offenen Monats anklicken** → rechts erscheint der "
    "Tag-für-Tag-Verlauf des As-of-Vergleichs."
)

_MONTH_NAMES_DE = ("Januar", "Februar", "März", "April", "Mai", "Juni", "Juli",
                   "August", "September", "Oktober", "November", "Dezember")


@st.cache_data(ttl=3600, show_spinner=False, max_entries=8)
def _monthly_asof(sig: str, year_new: int, year_old: int,
                  asof_n: pd.Timestamp, asof_o: pd.Timestamp,
                  _nig: pd.DataFrame) -> pd.DataFrame:
    """OTB je Stay-Monat am jeweiligen Stichtag (As-of, No-Shows raus).

    Bewusst NICHT über ``H.pace_by_month`` gebaut (Helper bleibt unangetastet):
    dessen Vorjahres-Stichtag ist fest auf −1 Jahr gespiegelt, hier gilt der
    frei gewählte ``asof_old`` (beliebiges Vergleichsjahr). Rechenkern ist
    dieselbe ``H.asof_on_the_books_mask``.
    """
    df = _nig
    if "is_no_show" in df.columns:
        df = df[~df["is_no_show"].astype(bool)]
    if df is None or df.empty:
        return pd.DataFrame()
    stay = pd.to_datetime(df["stay_date"]).dt.normalize()
    month = stay.dt.month
    rev = df["revenue"].astype(float)
    on_new = H.asof_on_the_books_mask(df, asof_n, include_cancellations=False)
    on_old = H.asof_on_the_books_mask(df, asof_o, include_cancellations=False)
    m_new = rev[on_new & (stay.dt.year == year_new)].groupby(month[on_new & (stay.dt.year == year_new)]).sum()
    m_old = rev[on_old & (stay.dt.year == year_old)].groupby(month[on_old & (stay.dt.year == year_old)]).sum()
    rows = []
    for m in range(1, 13):
        v_new = float(m_new.get(m, 0.0))
        v_old = float(m_old.get(m, 0.0))
        if v_new == 0.0 and v_old == 0.0:
            continue
        month_end = pd.Timestamp(year=year_new, month=m, day=1) + pd.offsets.MonthEnd(0)
        rows.append({
            "month_num": m,
            "Monat": _MONTH_NAMES_DE[m - 1],
            f"OTB {year_old} (€)": round(v_old, 0),
            f"OTB {year_new} (€)": round(v_new, 0),
            "Δ (€)": round(v_new - v_old, 0),
            "Δ (%)": round((v_new / v_old - 1) * 100.0, 1) if v_old > 0 else None,
            "offen": bool(month_end >= asof_n.normalize()),
        })
    return pd.DataFrame(rows)


@st.cache_data(ttl=3600, show_spinner=False, max_entries=16)
def _asof_evolution(sig: str, month: int, year_new: int, year_old: int,
                    year_delta: int, asof_n: pd.Timestamp,
                    _nig: pd.DataFrame, days_back: int = 90) -> pd.DataFrame:
    """Tages-Verlauf des As-of-OTB für EINEN Stay-Monat (NEW vs. gespiegeltes OLD).

    Für jeden As-of-Tag d (letzte ``days_back`` Tage bis zum Stichtag) wird der
    OTB des Monats mit ``H.asof_on_the_books_mask`` rekonstruiert; der
    OLD-Vergleichstag ist tagesgenau um ``year_delta`` Jahre gespiegelt.
    """
    df = _nig
    if "is_no_show" in df.columns:
        df = df[~df["is_no_show"].astype(bool)]
    if df is None or df.empty:
        return pd.DataFrame()
    stay = pd.to_datetime(df["stay_date"]).dt.normalize()
    sub_n = df[(stay.dt.year == year_new) & (stay.dt.month == month)]
    sub_o = df[(stay.dt.year == year_old) & (stay.dt.month == month)]
    days = pd.date_range(end=pd.Timestamp(asof_n).normalize(), periods=int(days_back), freq="D")
    rows = []
    for d in days:
        d_old = H.mirror_years(d, year_delta)
        v_n = (float(sub_n.loc[H.asof_on_the_books_mask(sub_n, d, include_cancellations=False),
                               "revenue"].sum()) if len(sub_n) else 0.0)
        v_o = (float(sub_o.loc[H.asof_on_the_books_mask(sub_o, d_old, include_cancellations=False),
                               "revenue"].sum()) if len(sub_o) else 0.0)
        rows.append({
            "date": d,
            "otb_new": v_n,
            "otb_old": v_o,
            "delta_abs": v_n - v_o,
            "delta_pct": (v_n / v_o - 1) * 100.0 if v_o > 0 else float("nan"),
        })
    return pd.DataFrame(rows)


def _evolution_fig(ev: pd.DataFrame, month_name: str, year_new: int, year_old: int):
    """Plotly-Figur: absolute OTB-Linien + Wedge (oben), Δ % je Tag (unten)."""
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots

    from revenueblindspots.theming import color as _bc

    fig = make_subplots(rows=2, cols=1, shared_xaxes=True,
                        row_heights=[0.66, 0.34], vertical_spacing=0.07)
    fig.add_trace(go.Scatter(
        x=ev["date"], y=ev["otb_old"], name=f"OTB {year_old} (gespiegelt)",
        line=dict(color="#666666", width=2, dash="dot"),
        hovertemplate="%{y:,.0f} €<extra>" + str(year_old) + "</extra>",
    ), row=1, col=1)
    fig.add_trace(go.Scatter(
        x=ev["date"], y=ev["otb_new"], name=f"OTB {year_new}",
        line=dict(color=_bc("blue"), width=2.6),
        fill="tonexty", fillcolor="rgba(255, 230, 80, 0.35)",  # Brand-Yellow-Wedge = Δ absolut
        hovertemplate="%{y:,.0f} €<extra>" + str(year_new) + "</extra>",
    ), row=1, col=1)
    _cols = [(_bc("green") if (pd.notna(v) and v >= 0) else _bc("red")) for v in ev["delta_pct"]]
    fig.add_trace(go.Bar(
        x=ev["date"], y=ev["delta_pct"], name="Δ vs Vorjahr (%)",
        marker_color=_cols, customdata=ev["delta_abs"],
        hovertemplate="Δ %{y:+.1f} %<br>Δ %{customdata:+,.0f} €<extra></extra>",
    ), row=2, col=1)
    fig.add_hline(y=0, line_width=1, line_color="#000000", row=2, col=1)
    fig.update_yaxes(title_text="OTB (€)", row=1, col=1, gridcolor="#EEEEEE", zeroline=False)
    fig.update_yaxes(title_text="Δ %", row=2, col=1, gridcolor="#EEEEEE", zeroline=False)
    fig.update_xaxes(gridcolor="#EEEEEE")
    fig.update_layout(
        height=520,
        paper_bgcolor="#FFFFFF", plot_bgcolor="#FFFFFF",
        font=dict(family="Neue Haas Grotesk Display Pro, Helvetica Neue, Helvetica, Arial, sans-serif",
                  color="#000000", size=13),
        hovermode="x unified",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, x=0),
        margin=dict(l=55, r=15, t=30, b=40),
        title=dict(text=f"{month_name} {year_new} · As-of-Verlauf (Tag für Tag)",
                   font=dict(size=15)),
    )
    return fig


_masof_sig = (
    f"{CD.snapshot_tag()}::{'+'.join(sorted(props_pick))}"
    f"::{asof_new.date()}::{asof_old.date()}::{YEAR_NEW}::{YEAR_OLD}"
)
_masof = _monthly_asof(_masof_sig, YEAR_NEW, YEAR_OLD, asof_new, asof_old, nightly)

if _masof.empty:
    alert_card("Keine OTB-Daten für die Monats-Sicht.", kind="info")
else:
    _mcolcfg = {
        "month_num": None,
        "Monat": st.column_config.TextColumn("Monat"),
        f"OTB {YEAR_OLD} (€)": st.column_config.NumberColumn(format="localized"),
        f"OTB {YEAR_NEW} (€)": st.column_config.NumberColumn(format="localized"),
        "Δ (€)": st.column_config.NumberColumn(format="localized"),
        "Δ (%)": st.column_config.NumberColumn(format="%+.1f %%"),
        "offen": st.column_config.CheckboxColumn(
            "offen?", help="Monatsende liegt am/nach dem Stichtag - der Monat füllt sich noch."
        ),
    }

    def _sel_rows(ev) -> list[int]:
        try:
            return list(ev.selection.rows)
        except Exception:
            try:
                return list(ev["selection"]["rows"])
            except Exception:
                return []

    _m_has_sel = bool(st.session_state.get("_pu_month_has_sel", False))
    if _m_has_sel:
        _mtcol, _mccol = st.columns([2, 3], gap="large")
    else:
        _mtcol, _mccol = st.container(), None

    with _mtcol:
        try:
            _mev = st.dataframe(
                _masof, hide_index=True, use_container_width=True,
                key="_pu_month_tbl", on_select="rerun", selection_mode="single-row",
                column_config=_mcolcfg,
            )
        except Exception:
            # Fallback ohne column_config (ältere Streamlit-Versionen).
            _mev = st.dataframe(
                _masof, hide_index=True, use_container_width=True,
                key="_pu_month_tbl_fb", on_select="rerun", selection_mode="single-row",
            )

    _mrows = _sel_rows(_mev)
    _sel_month = (
        int(_masof.iloc[_mrows[0]]["month_num"])
        if _mrows and 0 <= _mrows[0] < len(_masof)
        else None
    )
    _new_m_has = _sel_month is not None
    if _new_m_has != _m_has_sel:
        st.session_state["_pu_month_has_sel"] = _new_m_has
        st.rerun()

    if _sel_month is not None and _mccol is not None:
        _mrow = _masof[_masof["month_num"] == _sel_month].iloc[0]
        _mname = str(_mrow["Monat"])
        if bool(_mrow["offen"]):
            _ev_df = _asof_evolution(
                _masof_sig, _sel_month, YEAR_NEW, YEAR_OLD, YEAR_DELTA, asof_new, nightly
            )
            if _ev_df.empty:
                _mccol.info("Kein Tages-Verlauf verfügbar (keine Buchungen).")
            else:
                _last = _ev_df.iloc[-1]
                _mccol.plotly_chart(
                    _evolution_fig(_ev_df, _mname, YEAR_NEW, YEAR_OLD),
                    use_container_width=True, config={"displaylogo": False},
                )
                _mccol.caption(
                    f"**Heute ({asof_new:%d.%m.%Y}):** OTB {YEAR_NEW} "
                    f"{H.fmt_eur(float(_last['otb_new']))} vs. {YEAR_OLD} "
                    f"{H.fmt_eur(float(_last['otb_old']))} → "
                    f"Δ {float(_last['delta_abs']):+,.0f} € "
                    f"({float(_last['delta_pct']):+.1f} %). "
                    "Gelber Keil = Δ absolut; Balken unten = Δ relativ je As-of-Tag."
                    .replace(",", ".")
                )
        else:
            _mccol.info(
                f"**{_mname} {YEAR_NEW}** liegt vollständig vor dem Stichtag - "
                "der As-of-Wert ist final. Den Tag-für-Tag-Aufbau gibt es für "
                "offene/zukünftige Monate."
            )


# ============================== Downloads (3 separate) =====================
def _numeric_pickup_frame(
    raw: pd.DataFrame, label_col: str, rev_new_col: str, rev_old_col: str
) -> pd.DataFrame:
    """Pickup-Tabelle mit ROHEN Zahlen für den Excel-Export.

    Die Bildschirm-Tabellen (``disp_*``) enthalten formatierte Strings
    („1.234 €", „🟢 +5,2 %") - in Excel weder sortier- noch rechenbar.
    Hier: gleiche Kennzahlen als echte Zahlen (Review A12.10).
    """
    if raw is None or raw.empty:
        return pd.DataFrame()
    d = raw.copy()
    out = pd.DataFrame({
        label_col: d[label_col],
        f"Erstellt {YEAR_NEW} (EUR)": d[rev_new_col].astype(float).round(2),
        f"Erstellt {YEAR_OLD} (EUR)": d[rev_old_col].astype(float).round(2),
        "Delta absolut (EUR)": d["d_eur"].astype(float).round(2),
        "Delta relativ (%)": d["d_pct"].astype(float).round(1),
        f"OTB {YEAR_NEW} (EUR)": d["stay_new"].astype(float).round(2),
        f"OTB {YEAR_OLD} (EUR)": d["stay_old"].astype(float).round(2),
    })
    pu_new = GT._pickup_pct(d[rev_new_col], d["stay_new"])
    pu_old = GT._pickup_pct(d[rev_old_col], d["stay_old"])
    out[f"Pickup-Anteil {YEAR_NEW} (%)"] = pu_new.astype(float).round(1).to_numpy()
    out[f"Pickup-Anteil {YEAR_OLD} (%)"] = pu_old.astype(float).round(1).to_numpy()
    out["Delta Pickup (pp)"] = (pu_new - pu_old).astype(float).round(1).to_numpy()
    if "d_share_pp" in d.columns:
        out["Delta Anteil (pp)"] = d["d_share_pp"].astype(float).round(2)
    return out


def _tables_xlsx() -> bytes:
    """Aggregierte Pickup-Tabellen als Excel (3 Blätter, rohe Zahlen)."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    frames = [
        ("Standort", _numeric_pickup_frame(raw_loc, "Standort", "ist_new", "ist_old")),
        ("Buchungskanal", _numeric_pickup_frame(raw_ch, "Channel", "rev_new", "rev_old")),
        ("Stay-Segment", _numeric_pickup_frame(raw_seg, "Segment", "rev_new", "rev_old")),
    ]
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as xw:
        for name, df in frames:
            out = df if (df is not None and len(df)) else pd.DataFrame({"Hinweis": ["keine Daten"]})
            out.to_excel(xw, sheet_name=name, index=False)
        for ws in xw.book.worksheets:
            ws.freeze_panes = "A2"
            for i in range(1, ws.max_column + 1):
                ws.cell(row=1, column=i).font = Font(bold=True)
                ws.column_dimensions[get_column_letter(i)].width = 20
    return bio.getvalue()


def _raw_xlsx() -> bytes:
    """Roh-Timeslices (eine Zeile je Nacht, alle Flags)."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    frames = GT.stay_created_export_frames(
        nightly, props_pick, start_new, end_new, start_old, end_old,
        cre_start_new, cre_end_new, cre_start_old, cre_end_old,
        asof_new, asof_old, CMODE == CancelMode.ALL_IN,
    )
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl", datetime_format="YYYY-MM-DD HH:MM") as xw:
        for name, df in frames.items():
            out = df if len(df) else pd.DataFrame({"Hinweis": ["keine Daten in diesem Fenster"]})
            out.to_excel(xw, sheet_name=name, index=False)
        for ws in xw.book.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for i in range(1, ws.max_column + 1):
                ws.cell(row=1, column=i).font = Font(bold=True)
                ws.column_dimensions[get_column_letter(i)].width = 17
    return bio.getvalue()


def _curve_csv() -> bytes:
    """Buchungskurve-Daten (je Erstellungs-Tag, beide Jahre) als CSV."""
    if _line_df is None or _line_df.empty:
        return b"keine Daten"
    out = _line_df.copy()
    out["cum_rev_new"] = out["rev_new"].cumsum()
    out["cum_rev_old"] = out["rev_old"].cumsum()
    out[f"pickup_pct_{YEAR_NEW}"] = out["cum_rev_new"] / _otb_new * 100.0 if _otb_new else 0.0
    out[f"pickup_pct_{YEAR_OLD}"] = out["cum_rev_old"] / _otb_old * 100.0 if _otb_old else 0.0
    return out.to_csv(index=False).encode("utf-8")


st.markdown("## 6 · Downloads")
st.markdown("*Drei getrennte Exporte - genau für diese Sicht.*")
st.caption(
    "**Pickup-Tabellen (Excel)** = die drei aggregierten Tabellen oben, fertig "
    "fürs Deck. **Roh-Timeslices (Excel)** = eine Zeile je Nacht mit allen Flags "
    "zum Selber-Filtern. **Buchungskurve (CSV)** = je Erstellungs-Tag Umsatz + "
    "kumulierter Pickup-Anteil, beide Jahre."
)
d1, d2, d3 = st.columns(3)
_fname = f"pickup_{start_new:%Y%m%d}_vs_{start_old:%Y%m%d}"
with d1:
    st.download_button(
        "Pickup-Tabellen (Excel)",
        data=_tables_xlsx(),
        file_name=f"{_fname}_tabellen.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with d2:
    st.download_button(
        "Roh-Timeslices (Excel)",
        data=_raw_xlsx(),
        file_name=f"{_fname}_timeslices.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with d3:
    st.download_button(
        "Buchungskurve (CSV)",
        data=_curve_csv(),
        file_name=f"{_fname}_kurve.csv",
        mime="text/csv",
        use_container_width=True,
    )
