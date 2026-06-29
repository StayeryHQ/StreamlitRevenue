"""Roh-Export der reservations_timeslices als Excel (für Case-Studies / manuelles Filtern).

Liest den lokalen Snapshot (``data/timeslices.parquet``), filtert auf einen
Standort, ein Aufenthalts-Fenster (``serviceDate``) und ein Erstellungs-Fenster
(``created``) und schreibt die **rohen** Timeslice-Spalten (keine Transformation,
eine Zeile pro Stay-Nacht) in ein Excel-Sheet ``time_slices_download`` mit
aktivem AutoFilter und fixierter Kopfzeile.

Usage::

    python scripts/export_timeslices.py
    python scripts/export_timeslices.py --property CGN_WS \\
        --stay-start 2026-05-01 --stay-end 2026-07-12 \\
        --created-start 2026-03-01 --created-end 2026-07-02
    python scripts/export_timeslices.py --property BIE_HB --out /tmp/bie.xlsx

Hinweis: Es wird eine Zeile pro Stay-Nacht exportiert (so kommt
reservations_timeslices), NICHT pro Buchung. Werte bleiben unverändert.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd

_REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_REPO_ROOT / "src"))

from revenueblindspots import helpers as H  # noqa: E402

SHEET_NAME = "time_slices_download"

# Nur diese Spalten in den Export (alles andere raus). Hinweis: ``revenue`` und
# ``cancel_time`` (Stornodatum) sind engineered/broadcast - kein Roh-Timeslice-
# Feld -, werden hier aber bewusst ergänzt. ``vergleich`` (NEW/OLD) zuerst.
EXPORT_COLUMNS = [
    "vergleich",
    "bookingId",              # booking id
    "status",
    "created",                # creation date
    "cancel_time",            # cancellation date
    "serviceDate",            # Datum dieses Slices (die Nacht)
    "arrival",
    "departure",
    "revenue",                # = baseAmount_netAmount (engineered)
    "baseAmount_grossAmount",
    "baseAmount_netAmount",
]
# Nur zum Sortieren benötigt, fliegen vor dem Schreiben wieder raus.
_SORT_HELPERS = ["id", "serviceDate"]


def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    p.add_argument("--property", default="CGN_WS", help="Hotel-Code (default CGN_WS).")
    p.add_argument("--stay-start", default="2026-05-01", help="Aufenthalt von (serviceDate, inkl.).")
    p.add_argument("--stay-end", default="2026-07-12", help="Aufenthalt bis (serviceDate, inkl.).")
    p.add_argument("--created-start", default="2026-03-01", help="Erstellung von (created, inkl.).")
    p.add_argument("--created-end", default="2026-07-02", help="Erstellung bis (created, inkl.).")
    p.add_argument("--compare-offset", type=int, default=1,
                   help="Jahres-Versatz für das Vergleichsfenster (OLD). Default 1 "
                        "= Vorjahr gespiegelt. 0 = nur NEW, kein Vergleich.")
    p.add_argument("--snapshot-dir", default=None, help="Snapshot-Pfad. Default: data/ im Repo.")
    p.add_argument("--out", default=None, help="Ziel-xlsx. Default: <property>_timeslices_casestudy.xlsx im Repo.")
    return p.parse_args()


def _cohort(
    nig: pd.DataFrame,
    property_code: str,
    stay_start: pd.Timestamp,
    stay_end: pd.Timestamp,
    created_start: pd.Timestamp,
    created_end: pd.Timestamp,
    raw_cols: list[str],
    label: str,
) -> pd.DataFrame:
    """Roh-Timeslices eines Standorts, gefiltert auf Aufenthalts- + Erstellungs-Fenster."""
    sub = nig[nig["property_code"] == property_code]
    sub = H.filter_period(sub, stay_start, stay_end, "serviceDate")
    sub = H.filter_period(sub, created_start, created_end, "created")
    keep = list(dict.fromkeys(c for c in (raw_cols + _SORT_HELPERS) if c in sub.columns))
    sub = sub.loc[:, keep].copy()
    sub.insert(0, "vergleich", label)
    return sub


def export_timeslices(
    property_code: str,
    stay_start: str,
    stay_end: str,
    created_start: str,
    created_end: str,
    out_path: Path,
    compare_offset: int = 1,
    snapshot_dir: str | None = None,
) -> dict[str, int]:
    """Filtere die rohen Timeslices (NEW + gespiegeltes OLD) und schreibe sie als Excel.

    Gibt ein Dict ``{label: zeilen}`` zurück.
    """
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    nig = H.load_timeslices(snapshot_dir=snapshot_dir)
    # Nur die gewünschten Export-Spalten (ohne das später ergänzte ``vergleich``).
    raw_cols = [c for c in EXPORT_COLUMNS if c != "vergleich" and c in nig.columns]

    ss, se = pd.Timestamp(stay_start), pd.Timestamp(stay_end)
    cs, ce = pd.Timestamp(created_start), pd.Timestamp(created_end)

    frames = [_cohort(nig, property_code, ss, se, cs, ce, raw_cols, "NEW")]
    if compare_offset:
        # OLD-Fenster = NEW exakt um <compare_offset> Kalenderjahre gespiegelt.
        frames.append(
            _cohort(
                nig, property_code,
                H.mirror_years(ss, compare_offset), H.mirror_years(se, compare_offset),
                H.mirror_years(cs, compare_offset), H.mirror_years(ce, compare_offset),
                raw_cols, "OLD",
            )
        )

    sub = pd.concat(frames, ignore_index=True).sort_values(
        ["vergleich", "bookingId", "id", "serviceDate"]
    ).reset_index(drop=True)
    # Sortier-Hilfsspalten raus, nur die gewünschten Export-Spalten behalten.
    sub = sub[[c for c in EXPORT_COLUMNS if c in sub.columns]]

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl", datetime_format="YYYY-MM-DD HH:MM") as xw:
        sub.to_excel(xw, sheet_name=SHEET_NAME, index=False)
        ws = xw.sheets[SHEET_NAME]
        ws.auto_filter.ref = ws.dimensions          # AutoFilter über alle Spalten
        ws.freeze_panes = "A2"                       # Kopfzeile fixieren
        for i, col in enumerate(sub.columns, start=1):
            ws.cell(row=1, column=i).font = Font(bold=True)
            width = max(len(str(col)) + 2, 12)
            ws.column_dimensions[get_column_letter(i)].width = min(width, 40)
    return {lbl: int((sub["vergleich"] == lbl).sum()) for lbl in sub["vergleich"].unique()}


def main() -> None:
    args = _parse_args()
    out_path = (
        Path(args.out)
        if args.out
        else _REPO_ROOT / f"{args.property}_timeslices_casestudy.xlsx"
    )
    counts = export_timeslices(
        property_code=args.property,
        stay_start=args.stay_start,
        stay_end=args.stay_end,
        created_start=args.created_start,
        created_end=args.created_end,
        out_path=out_path,
        compare_offset=args.compare_offset,
        snapshot_dir=args.snapshot_dir,
    )
    total = sum(counts.values())
    print(f"[export] {total:,} Zeilen (Nächte) für {args.property} -> {out_path}")
    print(f"[export] davon {counts}")
    print(f"[export] NEW Aufenthalt {args.stay_start}..{args.stay_end} | "
          f"created {args.created_start}..{args.created_end} | "
          f"OLD = -{args.compare_offset} Jahr(e) gespiegelt")


if __name__ == "__main__":
    main()
